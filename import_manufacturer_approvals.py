"""
Import Airbus manufacturer/location approvals from a large XLSX file into PostgreSQL.

This script keeps only the fields we care about (part numbers, manufacturer,
location, and CAGE) while still preserving status metadata for troubleshooting.

Usage:
    python import_manufacturer_approvals.py /path/to/file.xlsx
"""
import argparse
import logging
import os
from datetime import datetime, date, timezone
from typing import Any, Dict, Iterator, List, Optional, Sequence, Tuple

from dotenv import load_dotenv
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_values

logger = logging.getLogger(__name__)

RAW_COLUMN_MAP = {
    "Manufacturer": "manufacturer_code",
    "Manufacturer Text": "manufacturer_name",
    "Location": "location",
    "Country": "country",
    "CAGE Co": "cage_code",
    "Approval Status Text": "approval_status",
    "Type of Data Text": "data_type",
    "Standard": "standard",
    "Airbus Material": "airbus_material",
    "Airbus Material Text": "airbus_material_text",
    "Interchangeability Flag": "interchangeability_flag",
    "Manufacturer PN": "manufacturer_part_number",
    "Usage Restriction Text": "usage_restriction",
    "P Status": "p_status",
    "P Status Text": "p_status_text",
    "Change date Status P": "status_change_date",
    "Counter of QIR": "qir_count",
}

UPSERT_COLUMNS: Sequence[str] = (
    "import_id",
    "manufacturer_code",
    "manufacturer_name",
    "location",
    "country",
    "cage_code",
    "approval_status",
    "data_type",
    "standard",
    "airbus_material",
    "airbus_material_text",
    "interchangeability_flag",
    "manufacturer_part_number",
    "usage_restriction",
    "p_status",
    "p_status_text",
    "status_change_date",
    "qir_count",
    "created_at",
    "updated_at",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Load Airbus manufacturer approvals into PostgreSQL."
    )
    parser.add_argument("xlsx_path", help="Path to the Airbus approvals XLSX export.")
    parser.add_argument(
        "--batch-size",
        type=int,
        default=2000,
        help="Number of rows to upsert per batch (default: 2000).",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limit number of data rows for testing or smoke runs.",
    )
    parser.add_argument(
        "--truncate-first",
        action="store_true",
        help="Remove existing approval rows before importing.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Run the parser and SQL without committing changes.",
    )
    return parser.parse_args()


def normalize_header(value: Any) -> str:
    return str(value).replace("\n", " ").strip()


def clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()
    text = str(value).strip()
    return text or None


def normalize_country(value: Any) -> Optional[str]:
    text = clean_text(value)
    return text.upper() if text else None


def normalize_cage(value: Any) -> Optional[str]:
    text = clean_text(value)
    return text.upper() if text else None


def parse_status_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    try:
        if len(text) == 8 and text.isdigit():
            return datetime.strptime(text, "%Y%m%d").date()
        return datetime.fromisoformat(text).date()
    except (ValueError, TypeError):
        logger.debug("Could not parse status date from %s", text)
        return None


def parse_int(value: Any) -> Optional[int]:
    text = clean_text(value)
    if text is None:
        return None
    try:
        return int(text)
    except (TypeError, ValueError):
        return None


def iter_records(xlsx_path: str, limit: Optional[int]) -> Iterator[Dict[str, Any]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [normalize_header(col) for col in header_row]

        records_yielded = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_map = dict(zip(headers, row))
            normalized: Dict[str, Any] = {}
            for raw_key, normalized_key in RAW_COLUMN_MAP.items():
                normalized[normalized_key] = row_map.get(raw_key)
            yield normalized
            records_yielded += 1
            if limit and records_yielded >= limit:
                break
    finally:
        workbook.close()


def build_payload(record: Dict[str, Any], import_id: int) -> Optional[Tuple[Any, ...]]:
    manufacturer_name = clean_text(record.get("manufacturer_name"))
    airbus_material = clean_text(record.get("airbus_material"))
    manufacturer_part_number = clean_text(record.get("manufacturer_part_number"))

    if not manufacturer_name:
        return None
    if not airbus_material and not manufacturer_part_number:
        return None

    now = datetime.now(tz=timezone.utc)
    return (
        import_id,
        clean_text(record.get("manufacturer_code")),
        manufacturer_name,
        clean_text(record.get("location")),
        normalize_country(record.get("country")),
        normalize_cage(record.get("cage_code")),
        clean_text(record.get("approval_status")),
        clean_text(record.get("data_type")),
        clean_text(record.get("standard")),
        airbus_material,
        clean_text(record.get("airbus_material_text")),
        clean_text(record.get("interchangeability_flag")),
        manufacturer_part_number,
        clean_text(record.get("usage_restriction")),
        clean_text(record.get("p_status")),
        clean_text(record.get("p_status_text")),
        parse_status_date(record.get("status_change_date")),
        parse_int(record.get("qir_count")),
        now,
        now,
    )


def start_import(cursor, source_file: str) -> int:
    cursor.execute(
        """
        INSERT INTO manufacturer_approval_imports (source_file, imported_by)
        VALUES (%s, %s)
        RETURNING id
        """,
        (os.path.basename(source_file), os.getenv("USER") or None),
    )
    return cursor.fetchone()[0]


def update_import_row_count(cursor, import_id: int, row_count: int) -> None:
    cursor.execute(
        "UPDATE manufacturer_approval_imports SET row_count = %s WHERE id = %s",
        (row_count, import_id),
    )


def upsert_batch(cursor, rows: List[Tuple[Any, ...]]) -> None:
    if not rows:
        return

    template = "(" + ", ".join(["%s"] * len(UPSERT_COLUMNS)) + ")"
    query = f"""
        INSERT INTO manufacturer_approvals ({", ".join(UPSERT_COLUMNS)})
        VALUES %s
        ON CONFLICT (airbus_material, manufacturer_part_number, manufacturer_name, cage_code, location)
        DO UPDATE SET
            import_id = EXCLUDED.import_id,
            country = COALESCE(EXCLUDED.country, manufacturer_approvals.country),
            approval_status = COALESCE(EXCLUDED.approval_status, manufacturer_approvals.approval_status),
            data_type = COALESCE(EXCLUDED.data_type, manufacturer_approvals.data_type),
            standard = COALESCE(EXCLUDED.standard, manufacturer_approvals.standard),
            airbus_material_text = COALESCE(EXCLUDED.airbus_material_text, manufacturer_approvals.airbus_material_text),
            interchangeability_flag = COALESCE(EXCLUDED.interchangeability_flag, manufacturer_approvals.interchangeability_flag),
            manufacturer_part_number = COALESCE(EXCLUDED.manufacturer_part_number, manufacturer_approvals.manufacturer_part_number),
            usage_restriction = COALESCE(EXCLUDED.usage_restriction, manufacturer_approvals.usage_restriction),
            p_status = COALESCE(EXCLUDED.p_status, manufacturer_approvals.p_status),
            p_status_text = COALESCE(EXCLUDED.p_status_text, manufacturer_approvals.p_status_text),
            status_change_date = COALESCE(EXCLUDED.status_change_date, manufacturer_approvals.status_change_date),
            qir_count = COALESCE(EXCLUDED.qir_count, manufacturer_approvals.qir_count),
            updated_at = EXCLUDED.updated_at
    """

    execute_values(cursor, query, rows, template=template, page_size=len(rows))


def process_file(
    connection,
    xlsx_path: str,
    batch_size: int,
    limit: Optional[int],
    truncate_first: bool,
    dry_run: bool,
) -> Dict[str, int]:
    cursor = connection.cursor()

    if truncate_first:
        logger.info("Truncating existing manufacturer approvals before import.")
        cursor.execute("TRUNCATE TABLE manufacturer_approvals RESTART IDENTITY")

    import_id = start_import(cursor, xlsx_path)
    stats = {"rows_seen": 0, "rows_written": 0, "rows_skipped": 0}
    batch: List[Tuple[Any, ...]] = []

    for record in iter_records(xlsx_path, limit):
        stats["rows_seen"] += 1
        payload = build_payload(record, import_id)
        if not payload:
            stats["rows_skipped"] += 1
            continue

        batch.append(payload)
        if len(batch) >= batch_size:
            upsert_batch(cursor, batch)
            stats["rows_written"] += len(batch)
            batch.clear()

    if batch:
        upsert_batch(cursor, batch)
        stats["rows_written"] += len(batch)

    update_import_row_count(cursor, import_id, stats["rows_written"])

    if dry_run:
        logger.info("Dry-run enabled; rolling back transaction.")
        connection.rollback()
    else:
        connection.commit()
        logger.info("Import committed.")

    logger.info(
        "Import summary: %s rows seen, %s written, %s skipped",
        stats["rows_seen"],
        stats["rows_written"],
        stats["rows_skipped"],
    )
    return stats


def main() -> None:
    args = parse_args()
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    load_dotenv()
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL is required to import manufacturer approvals.")

    connection = psycopg2.connect(database_url)
    try:
        process_file(
            connection,
            xlsx_path=args.xlsx_path,
            batch_size=args.batch_size,
            limit=args.limit,
            truncate_first=args.truncate_first,
            dry_run=args.dry_run,
        )
    finally:
        connection.close()


if __name__ == "__main__":
    main()
