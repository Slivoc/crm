"""
Airbus Marketplace routes
"""

from flask import Blueprint, jsonify, request, send_file, render_template
import calendar
import csv
from datetime import datetime, date
import io
import logging
import os
import time
import json
import re

from airbus_marketplace_helper import (
    suggest_marketplace_category,
    suggest_categories_batch,
    get_available_categories
)
from airbus_marketplace_export import export_parts_to_airbus_marketplace_csv
from models import get_db, convert_currency
from routes.portal_api import _analyze_quote_internal, get_portal_setting
from integrations.mirakl.client import MiraklClient, MiraklError
from integrations.mirakl.services.offers import build_offers_csv, OFFER_IMPORT_FIELDS
from openpyxl import load_workbook
from openai import OpenAI

logger = logging.getLogger(__name__)

marketplace_bp = Blueprint('marketplace', __name__)

PERPLEXITY_API_KEY = os.getenv('PERPLEXITY_API_KEY')
AIRBUS_ROTARY_APPROVAL_LIST_TYPE = 'airbus_rotary'
_AIRBUS_HARDWARE_REFERENCE_CACHE = None
_AIRBUS_HARDWARE_REFERENCE_CACHE_MTIME = None
_MASTER_LIST_TEST_REFERENCE_MAP = {
    '21215DC2405J': 'MKP-H-57077',
    'MS20470AD4-8': 'MKP-H-45486',
    'ASNA0045BC100L': 'MKP-H-59400',
}


def _months_ago(reference: date, months: int) -> date:
    if months <= 0:
        return reference

    year = reference.year
    month = reference.month - months
    while month <= 0:
        month += 12
        year -= 1

    day = min(reference.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def _get_mirakl_config():
    base_url = os.getenv('MIRAKL_BASE_URL') or get_portal_setting('mirakl_base_url')
    api_key = os.getenv('MIRAKL_API_KEY') or get_portal_setting('mirakl_api_key')
    shop_id = os.getenv('MIRAKL_SHOP_ID') or get_portal_setting('mirakl_shop_id')
    return base_url, api_key, shop_id


def _get_mirakl_client():
    base_url, api_key, shop_id = _get_mirakl_config()
    if not base_url or not api_key:
        return None, "MIRAKL_BASE_URL and MIRAKL_API_KEY must be configured."
    return MiraklClient(base_url, api_key, shop_id=shop_id), None


def _mask_api_key(value):
    if not value:
        return ''
    text = str(value)
    if len(text) <= 8:
        return '*' * len(text)
    return f"{text[:4]}...{text[-4:]}"


def _get_airbus_hardware_reference_workbook_path():
    docs_dir = os.path.join(os.getcwd(), 'docs')
    preferred = os.path.join(docs_dir, 'Copy of All Hardware References sept 2025.xlsx')
    if os.path.exists(preferred):
        return preferred
    return ''


def _upsert_portal_setting(cursor, key, value):
    cursor.execute(
        "UPDATE portal_settings SET setting_value = ?, date_modified = CURRENT_TIMESTAMP WHERE setting_key = ?",
        (value, key),
    )
    if cursor.rowcount == 0:
        cursor.execute(
            "INSERT INTO portal_settings (setting_key, setting_value, date_modified) VALUES (?, ?, CURRENT_TIMESTAMP)",
            (key, value),
        )


def _get_marketplace_customer_id():
    value = get_portal_setting('marketplace_customer_id')
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _get_perplexity_client():
    if not PERPLEXITY_API_KEY:
        return None
    return OpenAI(api_key=PERPLEXITY_API_KEY, base_url="https://api.perplexity.ai")


def _extract_json_object(raw_content):
    start = raw_content.find("{")
    end = raw_content.rfind("}")
    if start == -1 or end == -1 or end <= start:
        return None
    return raw_content[start:end + 1]


def _coerce_int(value, *, default=0):
    if value is None:
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    text = str(value).strip()
    if not text:
        return default
    try:
        return int(round(float(text.replace(',', '.'))))
    except ValueError:
        return default


def _coerce_price(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return ''
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip()
    if not text:
        return ''
    try:
        return round(float(text.replace(',', '.')), 2)
    except ValueError:
        return ''


def _convert_marketplace_price_to_eur(value):
    price = _coerce_price(value)
    if price == '':
        return ''
    try:
        return round(float(convert_currency(float(price), 'GBP', 'EUR')), 2)
    except Exception:
        logger.exception("Failed to convert marketplace price from GBP to EUR")
        return price


def _coerce_numeric_number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text.replace(',', '.'))
    except ValueError:
        return None


def _coerce_bool(value, *, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if text in ('1', 'true', 'yes', 'y', 'on'):
        return True
    if text in ('0', 'false', 'no', 'n', 'off'):
        return False
    return default


def _coerce_text(value, *, default=''):
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def _coerce_optional_text(value):
    text = _coerce_text(value, default='')
    return text if text else None


def _format_debug_number(value):
    number = _coerce_numeric_number(value)
    if number is None:
        return ''
    return round(number, 4)


def _build_offer_debug_row(part, offer_row, default_quantity):
    pricing_debug = part.get('pricing_debug') or {}
    source_details = pricing_debug.get('source_details') or {}
    winning_source = pricing_debug.get('winning_source') or part.get('price_source') or ''
    estimate_currency = part.get('source_currency') or 'GBP'

    raw_source_cost = source_details.get('cost')
    raw_source_currency = source_details.get('cost_currency') or estimate_currency
    source_cost_in_base = source_details.get('cost_in_base')
    margin_pct = source_details.get('margin_pct')
    target_price = source_details.get('target_price')
    rounded_price = source_details.get('rounded_price')
    exported_offer_price = offer_row.get('price')

    if raw_source_cost is None and source_cost_in_base is not None:
        raw_source_cost = source_cost_in_base
    if source_cost_in_base is None and raw_source_cost is not None and raw_source_currency == estimate_currency:
        source_cost_in_base = raw_source_cost
    if rounded_price is None:
        rounded_price = part.get('source_cost')

    pathway_steps = [f"winning_source {winning_source or 'unknown'}"]
    if raw_source_cost is not None:
        pathway_steps.append(f"raw_cost {round(float(raw_source_cost), 4)} {raw_source_currency}")
    if source_cost_in_base is not None:
        pathway_steps.append(f"base_cost {round(float(source_cost_in_base), 4)} {estimate_currency}")
    if margin_pct is not None:
        pathway_steps.append(f"margin {round(float(margin_pct), 4)}%")
    if target_price is not None:
        pathway_steps.append(f"target_price {round(float(target_price), 4)} {estimate_currency}")
    if rounded_price is not None:
        pathway_steps.append(f"rounded_price {round(float(rounded_price), 4)} {estimate_currency}")
    pathway_steps.append(f"offer_price {exported_offer_price}")

    return {
        'sku': offer_row.get('sku'),
        'product-id': offer_row.get('product-id'),
        'product-id-type': offer_row.get('product-id-type'),
        'winning_source': winning_source,
        'source_type': source_details.get('type') or '',
        'source_reference': source_details.get('reference') or '',
        'source_supplier': source_details.get('supplier') or '',
        'source_date': source_details.get('date') or '',
        'price_source': part.get('price_source') or '',
        'source_currency': raw_source_currency,
        'source_cost': _format_debug_number(raw_source_cost),
        'source_cost_in_gbp': _format_debug_number(source_cost_in_base),
        'margin_pct': _format_debug_number(margin_pct),
        'target_price_gbp': _format_debug_number(target_price),
        'rounded_price_gbp': _format_debug_number(rounded_price),
        'converted_price_eur': _format_debug_number(part.get('price')),
        'exported_offer_price': exported_offer_price,
        'in_stock': 'true' if part.get('source_in_stock') else 'false',
        'stock_qty': part.get('source_stock_qty'),
        'default_quantity_input': default_quantity,
        'exported_offer_quantity': offer_row.get('quantity'),
        'source_lead_days': '' if part.get('source_lead_days') is None else part.get('source_lead_days'),
        'exported_lead_days': offer_row.get('leadtime-to-ship'),
        'pricing_pathway': " -> ".join(pathway_steps),
    }


def _normalize_part_reference(value):
    return re.sub(r'[^A-Z0-9]+', '', str(value or '').strip().upper())


def _parse_reference_input(raw_value):
    return [
        item
        for item in (
            _coerce_text(chunk, default='')
            for chunk in re.split(r'[\r\n,;]+', str(raw_value or ''))
        )
        if item
    ]


def _iter_scalar_strings(value):
    if value is None:
        return
    if isinstance(value, dict):
        for item in value.values():
            yield from _iter_scalar_strings(item)
        return
    if isinstance(value, (list, tuple, set)):
        for item in value:
            yield from _iter_scalar_strings(item)
        return
    if isinstance(value, str):
        text = value.strip()
        if text:
            yield text


def _extract_mirakl_product_rows(payload):
    if isinstance(payload, dict):
        for key in ('products', 'data', 'items'):
            value = payload.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    return []


def _extract_mirakl_product_sku(product):
    for key in ('sku', 'product_sku', 'productSku', 'shop_sku', 'shopSku'):
        value = _coerce_optional_text(product.get(key))
        if value:
            return value
    return ''


def _extract_mirakl_product_label(product):
    for key in ('label', 'title', 'name', 'product', 'product_title', 'productTitle'):
        value = _coerce_optional_text(product.get(key))
        if value:
            return value
    return ''


def _match_mirakl_products_by_reference(products, references):
    remaining = {ref.upper(): ref for ref in references}
    matches = {}
    for product in products:
        scalar_values = {text.upper(): text for text in _iter_scalar_strings(product)}
        for ref_upper, original_ref in list(remaining.items()):
            if ref_upper in scalar_values:
                matches[original_ref] = product
                remaining.pop(ref_upper, None)
    return matches


def _collect_matching_candidates(products, reference):
    ref_upper = _coerce_text(reference, default='').upper()
    if not ref_upper:
        return []
    candidates = []
    for product in products:
        scalar_values = list(dict.fromkeys(text for text in _iter_scalar_strings(product)))
        matched_values = [text for text in scalar_values if text.upper() == ref_upper]
        if matched_values:
            candidates.append({
                'product': product,
                'matched_values': matched_values[:10],
                'scalar_preview': scalar_values[:25],
            })
    return candidates


def _normalize_marketplace_export_defaults(raw_defaults):
    payload = raw_defaults or {}
    description_mode = _coerce_text(payload.get('description_mode'), default='part_number')
    if description_mode not in ('part_number', 'none'):
        description_mode = 'part_number'

    package_content = _coerce_int(payload.get('package_content'), default=1)
    if package_content <= 0:
        package_content = 1

    return {
        'description_mode': description_mode,
        'mkp_eccn': _coerce_optional_text(payload.get('mkp_eccn')),
        'mkp_product_unit': _coerce_optional_text(payload.get('mkp_product_unit')),
        'mkp_package_content': package_content,
        'mkp_package_content_unit': _coerce_optional_text(payload.get('mkp_package_content_unit')),
        'mkp_third_level': _coerce_optional_text(payload.get('mkp_third_level')),
        'mkp_dangerous': _coerce_bool(payload.get('mkp_dangerous'), default=False),
        'mkp_serialized': _coerce_bool(payload.get('mkp_serialized'), default=False),
        'mkp_log_card': _coerce_bool(payload.get('mkp_log_card'), default=False),
        'mkp_easaf1': _coerce_bool(payload.get('mkp_easaf1'), default=False),
    }


def _get_marketplace_export_defaults():
    raw = {
        'description_mode': get_portal_setting('marketplace_export_default_description_mode'),
        'mkp_eccn': get_portal_setting('marketplace_export_default_mkp_eccn'),
        'mkp_product_unit': get_portal_setting('marketplace_export_default_mkp_product_unit'),
        'mkp_package_content': get_portal_setting('marketplace_export_default_mkp_package_content'),
        'mkp_package_content_unit': get_portal_setting('marketplace_export_default_mkp_package_content_unit'),
        'mkp_third_level': get_portal_setting('marketplace_export_default_mkp_third_level'),
        'mkp_dangerous': get_portal_setting('marketplace_export_default_mkp_dangerous'),
        'mkp_serialized': get_portal_setting('marketplace_export_default_mkp_serialized'),
        'mkp_log_card': get_portal_setting('marketplace_export_default_mkp_log_card'),
        'mkp_easaf1': get_portal_setting('marketplace_export_default_mkp_easaf1'),
    }
    return _normalize_marketplace_export_defaults(raw)


def _is_blank(value):
    if value is None:
        return True
    return isinstance(value, str) and not value.strip()


def _coerce_positive_number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip()
        if not text:
            return None
        try:
            number = float(text.replace(',', '.'))
        except ValueError:
            return None
    return number if number > 0 else None


def _part_identifier(record):
    return str(record.get('part_number') or record.get('base_part_number') or '').strip()


def _is_on_demand_without_price(record):
    commercial_mode = _coerce_text(record.get('commercial-on-collection'), default='').upper()
    return commercial_mode == 'ON_DEMAND' and _coerce_positive_number(record.get('price')) is None


def _get_offer_missing_required_fields(offer):
    missing = []
    if _is_blank(offer.get('sku')):
        missing.append('sku')
    if _is_blank(offer.get('product-id')):
        missing.append('product-id')
    if _is_blank(offer.get('product-id-type')):
        missing.append('product-id-type')
    if _coerce_numeric_number(offer.get('price')) is None:
        missing.append('price')
    if _coerce_positive_number(offer.get('quantity')) is None:
        missing.append('quantity')
    if _is_blank(offer.get('state')):
        missing.append('state')
    return missing


def _get_product_missing_required_fields(part):
    missing = []
    if not _part_identifier(part):
        missing.extend(['code', 'sku', 'product-id'])
    if _is_blank(part.get('mkp_category')):
        missing.append('mkpCategory')
    if _coerce_numeric_number(part.get('price')) is None:
        missing.append('price')
    if _coerce_positive_number(part.get('quantity')) is None:
        missing.append('quantity')
    if _is_blank(part.get('condition') or 'New'):
        missing.append('state')
    if _is_blank(part.get('mkp_description')) and _is_blank(part.get('description')):
        missing.append('description [en]')
    if _is_blank(part.get('mkp_eccn')):
        missing.append('eccn')
    if _is_blank(part.get('mkp_product_unit')):
        missing.append('productUnit')
    if _coerce_positive_number(part.get('mkp_package_content')) is None:
        missing.append('packageContent')
    if _is_blank(part.get('mkp_package_content_unit')):
        missing.append('packageContentUnit')
    if _is_blank(part.get('mkp_third_level')):
        missing.append('thirdLevel')
    if part.get('mkp_dangerous') is None:
        missing.append('dangerous')
    if part.get('mkp_serialized') is None:
        missing.append('serialized')
    if part.get('mkp_log_card') is None:
        missing.append('logCard')
    if part.get('mkp_easaf1') is None:
        missing.append('easaf1')
    return list(dict.fromkeys(missing))


def _split_valid_rows(rows, missing_fields_func, id_func):
    valid_rows = []
    invalid_rows = []
    for idx, row in enumerate(rows, start=1):
        missing = missing_fields_func(row)
        if missing:
            invalid_rows.append({
                'row': idx,
                'id': id_func(row),
                'missing_fields': missing,
            })
            continue
        valid_rows.append(row)
    return valid_rows, invalid_rows


def _dedupe_csv_headers(csv_bytes):
    text = csv_bytes.decode('utf-8-sig', errors='replace')
    if not text:
        return csv_bytes

    newline_index = text.find('\n')
    if newline_index == -1:
        header_line = text
        remainder = ''
        line_ending = '\n'
    else:
        header_line = text[:newline_index].rstrip('\r')
        remainder = text[newline_index + 1:]
        line_ending = '\n'

    if '\t' in header_line:
        delimiter = '\t'
    elif ',' in header_line:
        delimiter = ','
    elif ';' in header_line:
        delimiter = ';'
    else:
        return csv_bytes

    header = next(csv.reader([header_line], delimiter=delimiter))
    seen = {}
    updated_header = []
    changed = False
    for name in header:
        normalized = " ".join(str(name).split())
        count = seen.get(normalized, 0) + 1
        seen[normalized] = count
        if count > 1:
            updated_header.append(f"{normalized} {count}")
            changed = True
        else:
            updated_header.append(normalized)

    if not changed:
        return csv_bytes

    updated_line = delimiter.join(updated_header)
    rebuilt = f"{updated_line}{line_ending}{remainder}"
    return rebuilt.encode('utf-8')


def _normalize_import_header(header):
    return re.sub(r'[^a-z0-9]+', '', str(header or '').strip().lower())


def _parse_import_bool(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if not text:
        return None
    if text in ('1', 'true', 'yes', 'y', 'on'):
        return True
    if text in ('0', 'false', 'no', 'n', 'off'):
        return False
    return None


def _decode_csv_bytes(csv_bytes):
    try:
        return csv_bytes.decode('utf-8-sig')
    except UnicodeDecodeError:
        return csv_bytes.decode('latin-1', errors='replace')


def _detect_csv_delimiter(csv_text):
    first_line = csv_text.splitlines()[0] if csv_text else ''
    if '\t' in first_line:
        return '\t'
    if ';' in first_line and first_line.count(';') > first_line.count(','):
        return ';'
    return ','


def _read_import_rows(uploaded_file, filename):
    extension = os.path.splitext(filename.lower())[1]

    if extension == '.xlsx':
        workbook = load_workbook(io.BytesIO(uploaded_file.read()), read_only=True, data_only=True)
        worksheet = workbook.active
        row_iter = worksheet.iter_rows(values_only=True)
        headers = next(row_iter, None)
        if not headers:
            return []

        output_rows = []
        for row_number, values in enumerate(row_iter, start=2):
            row_data = {}
            for idx, raw_header in enumerate(headers):
                header = str(raw_header or '').strip()
                if not header:
                    continue
                row_data[header] = values[idx] if idx < len(values) else None
            output_rows.append((row_number, row_data))
        return output_rows

    csv_bytes = uploaded_file.read()
    if not csv_bytes:
        return []
    csv_text = _decode_csv_bytes(csv_bytes)
    delimiter = _detect_csv_delimiter(csv_text)
    reader = csv.DictReader(io.StringIO(csv_text), delimiter=delimiter)
    output_rows = []
    for row_number, row in enumerate(reader, start=2):
        output_rows.append((row_number, row))
    return output_rows


def _summarize_import_headers(rows):
    """Return raw/normalized header diagnostics for troubleshooting imports."""
    if not rows:
        return {'raw': [], 'normalized': []}

    first_row = rows[0][1] if len(rows[0]) > 1 else {}
    raw_headers = [str(header or '').strip() for header in (first_row or {}).keys()]
    normalized_headers = [_normalize_import_header(header) for header in raw_headers]
    return {
        'raw': raw_headers,
        'normalized': normalized_headers,
    }


def _detect_marketplace_import_template_kind(header_summary):
    normalized_headers = set((header_summary or {}).get('normalized') or [])
    if not normalized_headers:
        return 'unknown'

    product_headers = {
        'mkpcategory',
        'descriptionen',
        'productsummaryen',
        'productpresentationen',
        'productunit',
        'packagecontent',
        'packagecontentunit',
        'thirdlevel',
    }
    offer_headers = {
        'productid',
        'productidtype',
        'updatedelete',
        'allowquoterequests',
        'leadtimetoship',
        'vendorreference',
    }

    if normalized_headers.intersection(product_headers):
        return 'products'
    if normalized_headers.intersection(offer_headers):
        return 'offers'
    return 'unknown'


def _build_marketplace_updates_from_row(row_data, valid_categories):
    identifier_headers = {
        'code',
        'sku',
        'productid',
        'offersku',
        'productsku',
        'product',
        'mpn',
        'oem',
        'partnumber',
        'partno',
        'partnum',
        'manufacturerpn',
        'manufacturerpartnumber',
        'airbusmaterial',
    }

    # Airbus export column -> internal field mapping.
    column_map = {
        'code': 'part_number',
        'sku': 'part_number',
        'productid': 'part_number',
        # Mirakl offers export variants
        'offersku': 'part_number',
        'productsku': 'part_number',
        'product': 'part_number',
        'mpn': 'part_number',
        'oem': 'part_number',
        'partnumber': 'part_number',
        'partno': 'part_number',
        'partnum': 'part_number',
        'manufacturerpn': 'part_number',
        'manufacturerpartnumber': 'part_number',
        'airbusmaterial': 'part_number',
        'mkpcategory': 'mkp_category',
        'descriptionen': 'mkp_description',
        'name': 'mkp_name',
        'productsummaryen': 'mkp_product_summary',
        'productpresentationen': 'mkp_product_presentation',
        'productunit': 'mkp_product_unit',
        'packagecontent': 'mkp_package_content',
        'packagecontentunit': 'mkp_package_content_unit',
        'thirdlevel': 'mkp_third_level',
        'dangerous': 'mkp_dangerous',
        'eccn': 'mkp_eccn',
        'serialized': 'mkp_serialized',
        'logcard': 'mkp_log_card',
        'easaf1': 'mkp_easaf1',
        'id': 'mkp_offer_product_id',
    }

    extracted = {}
    identifier_candidates = []
    for raw_header, raw_value in (row_data or {}).items():
        normalized = _normalize_import_header(raw_header)
        mapped_field = column_map.get(normalized)
        if not mapped_field and normalized in identifier_headers:
            mapped_field = 'part_number'
        if not mapped_field:
            continue
        if mapped_field == 'part_number':
            identifier_candidates.append((normalized, raw_value))
            continue
        extracted[mapped_field] = raw_value

    # Choose identifier from best-to-worst columns so broad fields (e.g. Product)
    # do not override specific part references when both are present.
    identifier_preference = [
        'code',
        'sku',
        'productid',
        'offersku',
        'mpn',
        'oem',
        'partnumber',
        'partno',
        'partnum',
        'manufacturerpn',
        'manufacturerpartnumber',
        'airbusmaterial',
        'product',
        'productsku',
    ]
    for header_name in identifier_preference:
        for candidate_header, candidate_value in identifier_candidates:
            if candidate_header != header_name:
                continue
            candidate_text = str(candidate_value or '').strip()
            if candidate_text:
                extracted['part_number'] = candidate_text
                break
        if extracted.get('part_number'):
            break

    part_ref = str(extracted.get('part_number') or '').strip()
    if not part_ref:
        return None, None, (
            "Missing part identifier (expected one of: code, sku, product-id, offer sku,"
            " product, mpn, oem, part number, manufacturer pn, airbus material)."
        )

    updates = {}
    text_fields = (
        'mkp_description',
        'mkp_name',
        'mkp_product_summary',
        'mkp_product_presentation',
        'mkp_product_unit',
        'mkp_package_content_unit',
        'mkp_third_level',
        'mkp_eccn',
    )
    bool_fields = ('mkp_dangerous', 'mkp_serialized', 'mkp_log_card', 'mkp_easaf1')

    category_value = extracted.get('mkp_category')
    if category_value is not None:
        category_text = str(category_value).strip()
        if category_text:
            if category_text not in valid_categories:
                return part_ref, None, f"Invalid category: {category_text}"
            updates['mkp_category'] = category_text

    for field in text_fields:
        value = extracted.get(field)
        if value is None:
            continue
        value_text = str(value).strip()
        if value_text:
            updates[field] = value_text

    offer_product_id_value = extracted.get('mkp_offer_product_id')
    if offer_product_id_value is not None:
        offer_product_id_text = str(offer_product_id_value).strip()
        if offer_product_id_text:
            updates['mkp_offer_product_id'] = offer_product_id_text
            updates['mkp_offer_product_id_type'] = 'SKU'

    package_content_value = extracted.get('mkp_package_content')
    if package_content_value is not None:
        package_content_text = str(package_content_value).strip()
        if package_content_text:
            parsed = _coerce_int(package_content_text, default=None)
            if parsed is None:
                return part_ref, None, f"Invalid packageContent value: {package_content_text}"
            updates['mkp_package_content'] = parsed

    for field in bool_fields:
        parsed_bool = _parse_import_bool(extracted.get(field))
        if parsed_bool is not None:
            updates[field] = parsed_bool

    if not updates:
        return part_ref, None, "No importable marketplace detail fields found in row."

    return part_ref, updates, None


def _build_marketplace_baseline_from_row(row_data):
    identifier_headers = {
        'code', 'sku', 'productid', 'offersku', 'productsku', 'product', 'mpn', 'oem',
        'partnumber', 'partno', 'partnum', 'manufacturerpn', 'manufacturerpartnumber',
        'airbusmaterial',
    }

    identifier_candidates = []
    for raw_header, raw_value in (row_data or {}).items():
        normalized = _normalize_import_header(raw_header)
        if normalized in identifier_headers:
            identifier_candidates.append((normalized, raw_value))

    # Prefer seller-controlled identifiers first. In Airbus/Mirakl exports `code`
    # can be operator-side and not reliably match CRM part numbers.
    identifier_preference = [
        'sku', 'productid', 'offersku', 'productsku', 'product', 'mpn', 'oem',
        'partnumber', 'partno', 'partnum', 'manufacturerpn', 'manufacturerpartnumber',
        'airbusmaterial', 'code',
    ]
    part_ref = ''
    for header_name in identifier_preference:
        for candidate_header, candidate_value in identifier_candidates:
            if candidate_header != header_name:
                continue
            candidate_text = str(candidate_value or '').strip()
            if candidate_text:
                part_ref = candidate_text
                break
        if part_ref:
            break

    if not part_ref:
        return None, None, "Missing part identifier."

    baseline_row = {}
    for raw_header, raw_value in (row_data or {}).items():
        header = str(raw_header or '').strip()
        if not header:
            continue
        baseline_row[header] = raw_value

    return part_ref, baseline_row, None


def _get_import_row_value(row_data, target_header):
    target_normalized = _normalize_import_header(target_header)
    for raw_header, raw_value in (row_data or {}).items():
        if _normalize_import_header(raw_header) == target_normalized:
            return True, raw_value
    return False, None


def _extract_offer_product_identity(row_data):
    found_product_id, product_id = _get_import_row_value(row_data, 'product-id')
    product_id_text = _coerce_optional_text(product_id) if found_product_id else None
    product_id_type_text = None

    if product_id_text:
        _, product_id_type = _get_import_row_value(row_data, 'product-id-type')
        product_id_type_text = _coerce_optional_text(product_id_type) or 'SKU'
    else:
        found_product_sku, product_sku = _get_import_row_value(row_data, 'productsku')
        product_sku_text = _coerce_optional_text(product_sku) if found_product_sku else None
        if not product_sku_text:
            found_product_sku, product_sku = _get_import_row_value(row_data, 'product sku')
            product_sku_text = _coerce_optional_text(product_sku) if found_product_sku else None
        if not product_sku_text:
            found_shop_sku, shop_sku = _get_import_row_value(row_data, 'shop_sku')
            product_sku_text = _coerce_optional_text(shop_sku) if found_shop_sku else None
        if not product_sku_text:
            return None
        product_id_text = product_sku_text
        product_id_type_text = 'SKU'

    return {
        'product_id': product_id_text,
        'product_id_type': product_id_type_text,
    }


def _normalize_offer_product_id_type(product_id_type, product_id, part_number):
    normalized_type = _coerce_text(product_id_type, default='').strip()
    if normalized_type:
        return normalized_type

    product_id_text = _coerce_text(product_id, default='').strip()
    part_number_text = _coerce_text(part_number, default='').strip()
    if product_id_text and part_number_text and product_id_text == part_number_text:
        return 'mpnTitle'
    return 'SKU'


def _extract_product_identity(row_data):
    found_product_id, product_id = _get_import_row_value(row_data, 'id')
    product_id_text = _coerce_optional_text(product_id) if found_product_id else None
    if not product_id_text:
        return None

    return {
        'product_id': product_id_text,
        'product_id_type': 'SKU',
    }


def _resolve_offer_identity(part, source_mode):
    part_number = _part_identifier(part)
    baseline_row = part.get('baseline_row') or {}

    if source_mode == 'baseline':
        baseline_identity = _extract_offer_product_identity(baseline_row)
        if baseline_identity:
            baseline_product_id_text = _coerce_text(baseline_identity.get('product_id'), default='').strip()
            if baseline_product_id_text:
                return (
                    baseline_product_id_text,
                    _normalize_offer_product_id_type(
                        baseline_identity.get('product_id_type'),
                        baseline_product_id_text,
                        part_number,
                    ),
                )

    stored_product_id = _coerce_text(part.get('mkp_offer_product_id'), default='').strip()
    if stored_product_id:
        stored_product_id_type = _normalize_offer_product_id_type(
            part.get('mkp_offer_product_id_type'),
            stored_product_id,
            part_number,
        )
        return (stored_product_id, stored_product_id_type)

    resolved_mpn_title = _coerce_text(part.get('resolved_mpn_title'), default='').strip() or part_number
    return (resolved_mpn_title, 'mpnTitle')


def _sanitize_marketplace_lead_time_days(value, *, default=7):
    parsed = _coerce_int(value, default=default)
    if parsed is None:
        parsed = default
    return max(parsed, 1)


def _get_airbus_hardware_reference_maps():
    global _AIRBUS_HARDWARE_REFERENCE_CACHE, _AIRBUS_HARDWARE_REFERENCE_CACHE_MTIME

    workbook_path = _get_airbus_hardware_reference_workbook_path()
    if not workbook_path or not os.path.exists(workbook_path):
        return {
            'exact_titles': set(),
            'normalized_titles': set(),
            'alias_exact': {},
            'alias_normalized': {},
        }

    mtime = os.path.getmtime(workbook_path)
    if _AIRBUS_HARDWARE_REFERENCE_CACHE is not None and _AIRBUS_HARDWARE_REFERENCE_CACHE_MTIME == mtime:
        return _AIRBUS_HARDWARE_REFERENCE_CACHE

    exact_titles = set()
    normalized_titles = set()
    alias_exact_candidates = {}
    alias_normalized_candidates = {}

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        next(rows, None)

        for row in rows:
            raw_title = row[0] if len(row) > 0 else None
            raw_alts = row[1] if len(row) > 1 else None
            mpn_title = str(raw_title or '').strip()
            if not mpn_title:
                continue

            exact_titles.add(mpn_title.upper())
            normalized_title = _normalize_part_reference(mpn_title)
            if normalized_title:
                normalized_titles.add(normalized_title)

            alt_tokens = [
                token.strip()
                for token in str(raw_alts or '').split('|')
                if token and str(token).strip()
            ]
            for token in alt_tokens:
                token_upper = token.upper()
                alias_exact_candidates.setdefault(token_upper, set()).add(mpn_title)
                normalized_token = _normalize_part_reference(token)
                if normalized_token:
                    alias_normalized_candidates.setdefault(normalized_token, set()).add(mpn_title)
    finally:
        wb.close()

    alias_exact = {
        key: next(iter(values))
        for key, values in alias_exact_candidates.items()
        if len(values) == 1
    }
    alias_normalized = {
        key: next(iter(values))
        for key, values in alias_normalized_candidates.items()
        if len(values) == 1
    }

    _AIRBUS_HARDWARE_REFERENCE_CACHE = {
        'exact_titles': exact_titles,
        'normalized_titles': normalized_titles,
        'alias_exact': alias_exact,
        'alias_normalized': alias_normalized,
    }
    _AIRBUS_HARDWARE_REFERENCE_CACHE_MTIME = mtime
    return _AIRBUS_HARDWARE_REFERENCE_CACHE


def _resolve_airbus_mpn_title(part_number):
    part_text = _coerce_text(part_number, default='').strip()
    if not part_text:
        return {
            'resolved_mpn_title': '',
            'reference_status': 'missing',
            'reference_source': '',
        }

    maps = _get_airbus_hardware_reference_maps()
    part_upper = part_text.upper()
    normalized_part = _normalize_part_reference(part_text)

    if part_upper in maps['exact_titles']:
        return {
            'resolved_mpn_title': part_text,
            'reference_status': 'exact',
            'reference_source': 'hardware_refs_mpn_title',
        }
    if normalized_part and normalized_part in maps['normalized_titles']:
        return {
            'resolved_mpn_title': part_text,
            'reference_status': 'exact_normalized',
            'reference_source': 'hardware_refs_mpn_title',
        }

    alias_match = maps['alias_exact'].get(part_upper)
    if alias_match:
        return {
            'resolved_mpn_title': alias_match,
            'reference_status': 'alias',
            'reference_source': 'hardware_refs_alt_ref',
        }

    alias_normalized_match = maps['alias_normalized'].get(normalized_part) if normalized_part else None
    if alias_normalized_match:
        return {
            'resolved_mpn_title': alias_normalized_match,
            'reference_status': 'alias_normalized',
            'reference_source': 'hardware_refs_alt_ref',
        }

    return {
        'resolved_mpn_title': part_text,
        'reference_status': 'unknown',
        'reference_source': 'raw_part_number',
    }


_MARKETPLACE_MANUFACTURER_JOIN = """
    LEFT JOIN (
        SELECT
            pm.base_part_number,
            MIN(TRIM(m.name)) AS manufacturer_name
        FROM part_manufacturers pm
        JOIN manufacturers m ON m.id = pm.manufacturer_id
        WHERE TRIM(COALESCE(m.name, '')) <> ''
        GROUP BY pm.base_part_number
    ) marketplace_mfg ON marketplace_mfg.base_part_number = pn.base_part_number
"""


def _build_rotary_hqpl_exists_clause(part_alias='pn'):
    return f"""
        EXISTS (
            SELECT 1
            FROM manufacturer_approvals ma
            WHERE ma.approval_list_type = ?
              AND (
                  ma.airbus_material_base = {part_alias}.base_part_number
                  OR ma.manufacturer_part_number_base = {part_alias}.base_part_number
              )
        )
    """


def _get_global_alternative_map(cursor, base_part_numbers, rotary_only=True):
    entries_map = _get_global_alternative_entries_map(
        cursor,
        base_part_numbers,
        rotary_only=rotary_only,
    )

    return {
        source_base: [entry['part_number'] for entry in entries]
        for source_base, entries in entries_map.items()
    }


def _get_global_alternative_entries_map(cursor, base_part_numbers, rotary_only=True):
    cleaned = [str(value).strip() for value in (base_part_numbers or []) if str(value).strip()]
    if not cleaned:
        return {}

    placeholders = ','.join('?' * len(cleaned))
    query = f"""
        SELECT DISTINCT
            source.base_part_number AS source_base_part_number,
            alt.base_part_number AS alt_base_part_number,
            COALESCE(NULLIF(TRIM(pn_alt.part_number), ''), alt.base_part_number) AS alt_part_number
        FROM part_alt_group_members source
        JOIN part_alt_group_members alt
          ON alt.group_id = source.group_id
         AND alt.base_part_number <> source.base_part_number
        LEFT JOIN part_numbers pn_alt ON pn_alt.base_part_number = alt.base_part_number
        WHERE source.base_part_number IN ({placeholders})
    """
    params = list(cleaned)

    if rotary_only:
        query += """
          AND EXISTS (
              SELECT 1
              FROM manufacturer_approvals ma
              WHERE ma.approval_list_type = ?
                AND (
                    ma.airbus_material_base = alt.base_part_number
                    OR ma.manufacturer_part_number_base = alt.base_part_number
                )
          )
        """
        params.append(AIRBUS_ROTARY_APPROVAL_LIST_TYPE)

    query += """
        ORDER BY source.base_part_number, COALESCE(NULLIF(TRIM(pn_alt.part_number), ''), alt.base_part_number)
    """
    cursor.execute(query, params)
    rows = cursor.fetchall() or []

    alt_map = {}
    for row in rows:
        source_base = str(row['source_base_part_number']).strip()
        alt_base = str(row['alt_base_part_number'] or '').strip()
        alt_number = str(row['alt_part_number'] or '').strip()
        if not source_base or not alt_base or not alt_number:
            continue
        alt_map.setdefault(source_base, [])
        if not any(existing['base_part_number'] == alt_base for existing in alt_map[source_base]):
            alt_map[source_base].append({
                'base_part_number': alt_base,
                'part_number': alt_number,
            })

    return alt_map


def _get_alt_stock_rollup_map(customer_id, alt_entries_map):
    flat_requests = []
    for entries in (alt_entries_map or {}).values():
        for entry in entries:
            part_number = _coerce_text(entry.get('part_number'), default='')
            if not part_number:
                continue
            flat_requests.append({'part_number': part_number, 'quantity': 1})

    estimates = _get_portal_estimates(flat_requests, customer_id)
    rollup_map = {}

    for source_base_part_number, entries in (alt_entries_map or {}).items():
        total_stock_qty = 0
        highest_stock_price = None
        highest_stock_price_part_number = ''
        parts_with_stock = []

        for entry in entries:
            alt_base = _coerce_text(entry.get('base_part_number'), default='')
            alt_part_number = _coerce_text(entry.get('part_number'), default='')
            estimate = estimates.get(alt_base) or estimates.get(alt_part_number) or {}
            stock_qty = _coerce_int(estimate.get('stock_quantity'), default=0)
            in_stock = bool(estimate.get('in_stock')) and stock_qty > 0
            estimated_price = _coerce_numeric_number(estimate.get('estimated_price'))

            if not in_stock:
                continue

            total_stock_qty += stock_qty
            parts_with_stock.append(alt_part_number)

            if estimated_price is not None and (
                highest_stock_price is None or estimated_price > highest_stock_price
            ):
                highest_stock_price = estimated_price
                highest_stock_price_part_number = alt_part_number

        rollup_map[source_base_part_number] = {
            'has_alt_stock': total_stock_qty > 0,
            'stock_qty': total_stock_qty,
            'highest_stock_price': highest_stock_price,
            'highest_stock_price_part_number': highest_stock_price_part_number,
            'stock_part_numbers': parts_with_stock,
        }

    return rollup_map


def _select_alt_stock_rollup(include_non_hqpl_alts, rotary_rollup, all_rollup):
    return (all_rollup if include_non_hqpl_alts else rotary_rollup) or {}


def _apply_alt_stock_rollup_to_estimate(primary_estimate, alt_rollup):
    estimate = dict(primary_estimate or {})
    primary_stock_qty = _coerce_int(estimate.get('stock_quantity'), default=0)
    primary_in_stock = bool(estimate.get('in_stock')) and primary_stock_qty > 0
    primary_price = _coerce_numeric_number(estimate.get('estimated_price'))
    total_stock_qty = primary_stock_qty
    effective_price = primary_price
    price_source = estimate.get('price_source')

    alt_stock_qty = _coerce_int((alt_rollup or {}).get('stock_qty'), default=0)
    alt_highest_stock_price = _coerce_numeric_number((alt_rollup or {}).get('highest_stock_price'))
    if alt_stock_qty > 0:
        total_stock_qty += alt_stock_qty
        if alt_highest_stock_price is not None and (
            effective_price is None or alt_highest_stock_price > effective_price
        ):
            effective_price = alt_highest_stock_price
            price_source = 'alt_stock_rollup'

    if total_stock_qty > 0:
        estimate['in_stock'] = True
        estimate['stock_quantity'] = total_stock_qty
        estimate['estimated_lead_days'] = 0
        if effective_price is not None:
            estimate['estimated_price'] = effective_price
            estimate['price_source'] = price_source or estimate.get('price_source')

    return estimate


def _build_alt_stock_note(primary_stock_qty, alt_rollup):
    alt_rollup = alt_rollup or {}
    alt_stock_qty = _coerce_int(alt_rollup.get('stock_qty'), default=0)
    if alt_stock_qty <= 0:
        return ''

    alt_parts = [str(part).strip() for part in (alt_rollup.get('stock_part_numbers') or []) if str(part).strip()]
    if not alt_parts:
        return ''

    shown_parts = ', '.join(alt_parts[:3])
    suffix = f" (+{len(alt_parts) - 3} more)" if len(alt_parts) > 3 else ''
    if _coerce_int(primary_stock_qty, default=0) > 0:
        return f"Additional stock available via alternate PN(s): {shown_parts}{suffix}."
    return f"Stock held under alternate PN(s): {shown_parts}{suffix}."


def _build_offer_row_from_payload(offer):
    payload = {field: offer.get(field, '') for field in OFFER_IMPORT_FIELDS}
    baseline_row = offer.get('baseline_row') or {}

    for field in OFFER_IMPORT_FIELDS:
        found, value = _get_import_row_value(baseline_row, field)
        if found:
            payload[field] = value

    payload['price'] = offer.get('price', payload.get('price', ''))
    payload['quantity'] = offer.get('quantity', payload.get('quantity', ''))
    payload['sku'] = offer.get('sku', payload.get('sku', ''))
    payload['product-id'] = offer.get('product-id', payload.get('product-id', ''))
    payload['product-id-type'] = offer.get('product-id-type', payload.get('product-id-type', ''))
    payload['commercial-on-collection'] = offer.get(
        'commercial-on-collection',
        payload.get('commercial-on-collection', '')
    )

    if _is_on_demand_without_price(payload):
        payload['price'] = ''
        payload['price-additional-info'] = ''
        payload['discount-price'] = ''
        for field in OFFER_IMPORT_FIELDS:
            lowered = field.lower()
            if lowered.startswith('price[') or lowered.startswith('discount-price['):
                payload[field] = ''

    return payload


def _normalize_prefixes(prefixes):
    cleaned = []
    for prefix in prefixes or []:
        value = (prefix or "").strip()
        if value:
            cleaned.append(value)
    return list(dict.fromkeys(cleaned))


def _guess_prefixes_from_part_number(part_number):
    if not part_number:
        return []
    token = re.split(r"[-/\\s]+", part_number.strip())[0]
    if token and token != part_number:
        return [token]
    return []


def _build_prefix_conditions(prefixes):
    conditions = []
    params = []
    for prefix in prefixes:
        like_value = f"{prefix}%"
        conditions.append("(pn.part_number LIKE ? OR pn.base_part_number LIKE ?)")
        params.extend([like_value, like_value])
    return conditions, params


def _get_portal_estimates(parts, customer_id):
    if not parts:
        return {}

    estimate_response = _analyze_quote_internal(customer_id, parts)
    estimate_data = (
        estimate_response[0].get_json()
        if isinstance(estimate_response, tuple)
        else estimate_response.get_json()
    )
    if not estimate_data or not estimate_data.get('success'):
        logger.info(
            "Marketplace export estimates unavailable: customer_id=%s success=%s",
            customer_id,
            estimate_data.get('success') if estimate_data else None,
        )
        return {}

    results = estimate_data.get('results', [])
    return {
        (item.get('base_part_number') or item.get('part_number')): item
        for item in results
        if item.get('base_part_number') or item.get('part_number')
    }


def _fetch_marketplace_parts_by_references(references):
    cleaned_refs = [ref.strip() for ref in references if ref and ref.strip()]
    if not cleaned_refs:
        return []

    exact_references = {ref.upper() for ref in cleaned_refs}
    normalized_references = {
        normalized
        for normalized in (_normalize_part_reference(ref) for ref in cleaned_refs)
        if normalized
    }
    exact_placeholders = ','.join(['?'] * len(exact_references)) if exact_references else ''
    normalized_placeholders = ','.join(['?'] * len(normalized_references)) if normalized_references else ''
    where_clauses = []
    params = []

    if exact_references:
        where_clauses.append(f"UPPER(pn.base_part_number) IN ({exact_placeholders})")
        params.extend(exact_references)
        where_clauses.append(f"UPPER(pn.part_number) IN ({exact_placeholders})")
        params.extend(exact_references)

    if normalized_references:
        where_clauses.append(
            f"REGEXP_REPLACE(UPPER(COALESCE(pn.base_part_number, '')), '[^A-Z0-9]+', '', 'g') IN ({normalized_placeholders})"
        )
        params.extend(normalized_references)
        where_clauses.append(
            f"REGEXP_REPLACE(UPPER(COALESCE(pn.part_number, '')), '[^A-Z0-9]+', '', 'g') IN ({normalized_placeholders})"
        )
        params.extend(normalized_references)

    if not where_clauses:
        return []

    db = get_db()
    cursor = db.cursor()
    cursor.execute(
        f"""
        SELECT
            pn.base_part_number,
            pn.part_number,
            marketplace_mfg.manufacturer_name,
            pn.mkp_category,
            pn.mkp_description,
            pn.mkp_name,
            pn.mkp_offer_product_id,
            pn.mkp_offer_product_id_type
        FROM part_numbers pn
        {_MARKETPLACE_MANUFACTURER_JOIN}
        WHERE {' OR '.join(where_clauses)}
        ORDER BY pn.part_number
        """,
        params,
    )
    rows = cursor.fetchall()

    customer_id = _get_marketplace_customer_id()
    parts_payload = [
        {'part_number': row['part_number'] or row['base_part_number'], 'quantity': 1}
        for row in rows
    ]
    estimates = _get_portal_estimates(parts_payload, customer_id)

    exact_lookup = {}
    normalized_lookup = {}
    for row in rows:
        part_number = row['part_number'] or row['base_part_number']
        estimate = estimates.get(row['base_part_number']) or estimates.get(part_number) or {}
        part = {
            'base_part_number': row['base_part_number'],
            'part_number': part_number,
            'manufacturer': _coerce_text(row['manufacturer_name'], default=''),
            'mkp_category': row['mkp_category'],
            'mkp_description': row['mkp_description'],
            'mkp_name': row['mkp_name'],
            'mkp_offer_product_id': row['mkp_offer_product_id'],
            'mkp_offer_product_id_type': row['mkp_offer_product_id_type'],
            'estimated_price_gbp': estimate.get('estimated_price'),
            'estimated_price_eur': _convert_marketplace_price_to_eur(estimate.get('estimated_price')),
            'estimated_lead_days': estimate.get('estimated_lead_days'),
            'in_stock': bool(estimate.get('in_stock')),
            'stock_qty': estimate.get('stock_quantity') if estimate.get('stock_quantity') is not None else 0,
            'price_source': estimate.get('price_source'),
        }
        for value in (row['base_part_number'], row['part_number']):
            if value:
                exact_lookup.setdefault(str(value).strip().upper(), part)
                normalized = _normalize_part_reference(value)
                if normalized:
                    normalized_lookup.setdefault(normalized, part)

    ordered = []
    for ref in cleaned_refs:
        ref_key = ref.strip().upper()
        normalized_ref = _normalize_part_reference(ref)
        ordered.append({
            'requested_reference': ref,
            'crm_part': exact_lookup.get(ref_key) or (normalized_lookup.get(normalized_ref) if normalized_ref else None),
        })
    return ordered


def _store_offer_identity_updates_from_rows(rows):
    identity_by_reference = {}
    for row in rows or []:
        sku = _coerce_text(row.get('sku'), default='').strip()
        product_id = _coerce_text(row.get('product-id'), default='').strip()
        product_id_type = _coerce_text(row.get('product-id-type'), default='').strip()
        if not sku or not product_id or not product_id_type:
            continue
        identity_by_reference[sku.upper()] = {
            'normalized_reference': _normalize_part_reference(sku),
            'product_id': product_id,
            'product_id_type': product_id_type,
        }

    if not identity_by_reference:
        return 0

    exact_references = list(identity_by_reference.keys())
    normalized_references = list({
        payload['normalized_reference']
        for payload in identity_by_reference.values()
        if payload.get('normalized_reference')
    })
    exact_placeholders = ','.join(['?'] * len(exact_references)) if exact_references else ''
    normalized_placeholders = ','.join(['?'] * len(normalized_references)) if normalized_references else ''
    where_clauses = []
    params = []

    if exact_references:
        where_clauses.append(f"UPPER(base_part_number) IN ({exact_placeholders})")
        params.extend(exact_references)
        where_clauses.append(f"UPPER(part_number) IN ({exact_placeholders})")
        params.extend(exact_references)

    if normalized_references:
        where_clauses.append(
            f"REGEXP_REPLACE(UPPER(COALESCE(base_part_number, '')), '[^A-Z0-9]+', '', 'g') IN ({normalized_placeholders})"
        )
        params.extend(normalized_references)
        where_clauses.append(
            f"REGEXP_REPLACE(UPPER(COALESCE(part_number, '')), '[^A-Z0-9]+', '', 'g') IN ({normalized_placeholders})"
        )
        params.extend(normalized_references)

    if not where_clauses:
        return 0

    db = get_db()
    cursor = db.cursor()
    cursor.execute(
        f"""
        SELECT base_part_number, part_number
        FROM part_numbers
        WHERE {' OR '.join(where_clauses)}
        """,
        params,
    )
    matches = cursor.fetchall()

    updates = {}
    for match in matches:
        base_part_number = match['base_part_number']
        part_number = match['part_number']
        for value in (base_part_number, part_number):
            if not value:
                continue
            payload = identity_by_reference.get(str(value).strip().upper())
            if not payload:
                normalized_value = _normalize_part_reference(value)
                for candidate in identity_by_reference.values():
                    if normalized_value and candidate.get('normalized_reference') == normalized_value:
                        payload = candidate
                        break
            if payload:
                updates[str(base_part_number)] = (
                    payload['product_id'],
                    payload['product_id_type'],
                    base_part_number,
                )
                break

    if not updates:
        return 0

    cursor.executemany(
        """
        UPDATE part_numbers
        SET mkp_offer_product_id = ?, mkp_offer_product_id_type = ?
        WHERE base_part_number = ?
        """,
        list(updates.values()),
    )
    db.commit()
    return len(updates)


def _build_master_list_test_offers():
    ordered_refs = list(_MASTER_LIST_TEST_REFERENCE_MAP.keys())
    crm_rows = _fetch_marketplace_parts_by_references(ordered_refs)
    crm_lookup = {
        row['requested_reference']: row.get('crm_part')
        for row in crm_rows
    }

    offers = []
    skipped = []
    for reference in ordered_refs:
        crm_part = crm_lookup.get(reference)
        if not crm_part:
            skipped.append({'reference': reference, 'reason': 'CRM part not found'})
            continue

        price = crm_part.get('estimated_price_eur')
        quantity = crm_part.get('stock_qty') if crm_part.get('in_stock') else 1
        offer = _build_offer_row_from_payload({
            'sku': crm_part.get('part_number') or crm_part.get('base_part_number') or reference,
            'product-id': _MASTER_LIST_TEST_REFERENCE_MAP[reference],
            'product-id-type': 'SKU',
            'description': crm_part.get('mkp_description') or crm_part.get('part_number') or reference,
            'internal-description': '',
            'price': price,
            'price-additional-info': '',
            'quantity': quantity,
            'min-quantity-alert': '',
            'state': '1',
            'available-start-date': '',
            'available-end-date': '',
            'logistic-class': '',
            'favorite-rank': '',
            'discount-start-date': '',
            'discount-end-date': '',
            'discount-price': '',
            'update-delete': 'update',
            'allow-quote-requests': 'true',
            'leadtime-to-ship': _sanitize_marketplace_lead_time_days(
                crm_part.get('estimated_lead_days'),
                default=7,
            ),
            'min-order-quantity': '',
            'max-order-quantity': '',
            'package-quantity': '',
            'commercial-on-collection': 'ON_DEMAND',
            'plt': '',
            'plt-unit': '',
            'shelflife': '',
            'shelflife-unit': '',
            'warranty': '',
            'warranty-unit': '',
            'up-sell': '',
            'cross-sell': '',
            'vendor-reference': 'master-list-test',
        })
        missing = _get_offer_missing_required_fields(offer)
        if missing:
            skipped.append({
                'reference': reference,
                'reason': f"Missing required fields: {', '.join(missing)}",
            })
            continue
        offers.append(offer)

    if not offers:
        raise ValueError(
            'No valid master-list test offers could be built. '
            + '; '.join(f"{item['reference']}: {item['reason']}" for item in skipped[:10])
        )

    return offers, skipped


@marketplace_bp.route('/import-details-file', methods=['POST'])
def import_marketplace_details_file():
    """
    Import Airbus marketplace CSV/XLSX and update non-commercial marketplace fields.
    Price/quantity/stock values are intentionally ignored.
    """
    db = None
    try:
        file = request.files.get('file')
        if not file or not file.filename:
            return jsonify({'success': False, 'error': 'CSV/XLSX file is required'}), 400

        filename = file.filename
        extension = os.path.splitext(filename.lower())[1]
        if extension not in ('.csv', '.xlsx'):
            return jsonify({'success': False, 'error': 'Only .csv and .xlsx files are supported'}), 400

        rows = _read_import_rows(file, filename)
        if not rows:
            return jsonify({'success': False, 'error': 'Uploaded file is empty'}), 400

        header_summary = _summarize_import_headers(rows)
        total_rows = len(rows)
        logger.info("Marketplace details import started: filename=%s rows=%s", filename, total_rows)

        valid_categories = set(get_available_categories())
        db = get_db()
        cursor = db.cursor()

        processed = 0
        updated = 0
        changed = 0
        skipped_no_match = 0
        skipped_no_updates = 0
        row_errors = []
        error_reason_counts = {}

        def _record_error(reason):
            error_reason_counts[reason] = error_reason_counts.get(reason, 0) + 1

        for row_number, row_data in rows:
            processed += 1
            part_ref, field_updates, row_error = _build_marketplace_updates_from_row(row_data, valid_categories)
            if row_error:
                skipped_no_updates += 1
                _record_error(row_error)
                if len(row_errors) < 20:
                    row_errors.append(f"Row {row_number}: {row_error}")
                continue

            set_clause = ", ".join(f"{field} = ?" for field in field_updates.keys())
            part_ref_upper = part_ref.strip().upper()
            normalized_part_ref = _normalize_part_reference(part_ref)
            params = list(field_updates.values()) + [
                part_ref_upper,
                part_ref_upper,
                normalized_part_ref,
                normalized_part_ref,
            ]
            cursor.execute(
                f"""
                UPDATE part_numbers
                SET {set_clause}
                WHERE UPPER(base_part_number) = ?
                   OR UPPER(part_number) = ?
                   OR REGEXP_REPLACE(UPPER(COALESCE(base_part_number, '')), '[^A-Z0-9]+', '', 'g') = ?
                   OR REGEXP_REPLACE(UPPER(COALESCE(part_number, '')), '[^A-Z0-9]+', '', 'g') = ?
                """,
                params
            )
            if cursor.rowcount > 0:
                updated += cursor.rowcount
                changed += cursor.rowcount
            else:
                skipped_no_match += 1
                _record_error("Part not found in CRM")
                if len(row_errors) < 20:
                    row_errors.append(f"Row {row_number}: Part not found for '{part_ref}'.")

            if processed % 250 == 0:
                logger.info(
                    "Marketplace details import progress: processed=%s/%s matched=%s no_match=%s skipped=%s",
                    processed,
                    total_rows,
                    updated,
                    skipped_no_match,
                    skipped_no_updates,
                )

        db.commit()
        logger.info(
            "Marketplace details import completed: processed=%s matched=%s changed=%s no_match=%s skipped=%s",
            processed,
            updated,
            changed,
            skipped_no_match,
            skipped_no_updates,
        )

        return jsonify({
            'success': True,
            'message': 'Marketplace detail import complete.',
            'summary': {
                'processed_rows': processed,
                'updated_rows': updated,
                'changed_rows': changed,
                'skipped_no_match': skipped_no_match,
                'skipped_no_updates': skipped_no_updates,
            },
            'errors': row_errors,
            'error_reason_counts': error_reason_counts,
            'header_diagnostics': header_summary,
        }), 200
    except Exception as e:
        logger.exception("Error importing marketplace details file")
        if db:
            db.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@marketplace_bp.route('/import-stock-price-file', methods=['POST'])
def import_marketplace_stock_price_file():
    """
    Parse Airbus marketplace export CSV/XLSX and retain Airbus rows as an export baseline.
    This endpoint does not write to DB; it returns matched baseline rows for export/push,
    while CRM price and quantity remain authoritative.
    """
    try:
        file = request.files.get('file')
        if not file or not file.filename:
            return jsonify({'success': False, 'error': 'CSV/XLSX file is required'}), 400

        filename = file.filename
        extension = os.path.splitext(filename.lower())[1]
        if extension not in ('.csv', '.xlsx'):
            return jsonify({'success': False, 'error': 'Only .csv and .xlsx files are supported'}), 400

        rows = _read_import_rows(file, filename)
        if not rows:
            return jsonify({'success': False, 'error': 'Uploaded file is empty'}), 400

        header_summary = _summarize_import_headers(rows)
        template_kind = _detect_marketplace_import_template_kind(header_summary)
        db = get_db()
        cursor = db.cursor()

        processed = 0
        skipped_invalid = 0
        skipped_no_match = 0
        row_errors = []
        parsed_rows = []
        exact_references = set()
        normalized_references = set()
        offer_identity_updates = {}
        matched_offer_identity_count = 0

        for row_number, row_data in rows:
            processed += 1
            part_ref, baseline_row, row_error = _build_marketplace_baseline_from_row(row_data)
            if row_error:
                skipped_invalid += 1
                if len(row_errors) < 20:
                    row_errors.append(f"Row {row_number}: {row_error}")
                continue
            ref_key = part_ref.strip().upper()
            normalized_ref = _normalize_part_reference(part_ref)
            exact_references.add(ref_key)
            if normalized_ref:
                normalized_references.add(normalized_ref)
            parsed_rows.append((row_number, part_ref.strip(), ref_key, normalized_ref, baseline_row))

        if not exact_references and not normalized_references:
            return jsonify({
                'success': False,
                'error': 'No valid part rows found in file.',
                'errors': row_errors,
            }), 400

        exact_placeholders = ','.join(['?'] * len(exact_references)) if exact_references else ''
        normalized_placeholders = ','.join(['?'] * len(normalized_references)) if normalized_references else ''
        where_clauses = []
        params = []

        if exact_references:
            where_clauses.append(f"UPPER(base_part_number) IN ({exact_placeholders})")
            params.extend(exact_references)
            where_clauses.append(f"UPPER(part_number) IN ({exact_placeholders})")
            params.extend(exact_references)

        if normalized_references:
            where_clauses.append(
                f"REGEXP_REPLACE(UPPER(COALESCE(base_part_number, '')), '[^A-Z0-9]+', '', 'g') IN ({normalized_placeholders})"
            )
            params.extend(normalized_references)
            where_clauses.append(
                f"REGEXP_REPLACE(UPPER(COALESCE(part_number, '')), '[^A-Z0-9]+', '', 'g') IN ({normalized_placeholders})"
            )
            params.extend(normalized_references)

        cursor.execute(
            f"""
            SELECT base_part_number, part_number
            FROM part_numbers
            WHERE {' OR '.join(where_clauses)}
            """,
            params,
        )
        matches = cursor.fetchall()

        exact_lookup = {}
        normalized_lookup = {}
        for row in matches:
            base_part_number = row['base_part_number']
            part_number = row['part_number']
            matched_value = (base_part_number, part_number)
            if base_part_number:
                exact_lookup.setdefault(str(base_part_number).strip().upper(), matched_value)
                normalized_base = _normalize_part_reference(base_part_number)
                if normalized_base:
                    normalized_lookup.setdefault(normalized_base, matched_value)
            if part_number:
                exact_lookup.setdefault(str(part_number).strip().upper(), matched_value)
                normalized_part = _normalize_part_reference(part_number)
                if normalized_part:
                    normalized_lookup.setdefault(normalized_part, matched_value)

        matched_baselines = {}
        for row_number, part_ref, ref_key, normalized_ref, baseline_row in parsed_rows:
            matched = exact_lookup.get(ref_key)
            if not matched and normalized_ref:
                matched = normalized_lookup.get(normalized_ref)
            if not matched:
                skipped_no_match += 1
                if len(row_errors) < 20:
                    row_errors.append(f"Row {row_number}: Part not found for '{part_ref}'.")
                continue
            base_part_number, part_number = matched
            matched_baselines[str(base_part_number)] = {
                'base_part_number': base_part_number,
                'part_number': part_number or base_part_number,
                'baseline_row': baseline_row,
            }
            if template_kind == 'offers':
                offer_identity = _extract_offer_product_identity(baseline_row)
                if offer_identity:
                    offer_identity_updates[str(base_part_number)] = offer_identity
                    matched_offer_identity_count += 1
            elif template_kind == 'products':
                product_identity = _extract_product_identity(baseline_row)
                if product_identity:
                    offer_identity_updates[str(base_part_number)] = product_identity
                    matched_offer_identity_count += 1

        stored_offer_product_ids_count = 0
        if offer_identity_updates:
            cursor.executemany(
                """
                UPDATE part_numbers
                SET mkp_offer_product_id = ?, mkp_offer_product_id_type = ?
                WHERE base_part_number = ?
                """,
                [
                    (
                        identity['product_id'],
                        identity['product_id_type'],
                        base_part_number,
                    )
                    for base_part_number, identity in offer_identity_updates.items()
                ],
            )
            db.commit()
            stored_offer_product_ids_count = len(offer_identity_updates)

        template_label = {
            'offers': 'offer baseline',
            'products': 'product baseline',
        }.get(template_kind, 'Airbus baseline')

        return jsonify({
            'success': True,
            'message': (
                f'Airbus {template_label} imported. '
                'Export/push will keep uploaded Airbus fields and replace price/quantity from CRM.'
            ),
            'summary': {
                'processed_rows': processed,
                'matched_rows': len(matched_baselines),
                'skipped_invalid': skipped_invalid,
                'skipped_no_match': skipped_no_match,
                'matched_with_offer_identity': matched_offer_identity_count,
                'matched_missing_offer_identity': max(len(matched_baselines) - matched_offer_identity_count, 0),
            },
            'header_diagnostics': header_summary,
            'template_kind': template_kind,
            'stored_offer_product_ids_count': stored_offer_product_ids_count,
            'baselines': list(matched_baselines.values()),
            'errors': row_errors,
        }), 200
    except Exception as e:
        logger.exception("Error importing marketplace stock/price file")
        return jsonify({'success': False, 'error': str(e)}), 500


@marketplace_bp.route('/export-page', methods=['GET'])
def export_page():
    """Render the marketplace export page"""
    from flask import render_template
    mirakl_base_url = get_portal_setting('mirakl_base_url')
    mirakl_shop_id = get_portal_setting('mirakl_shop_id')
    mirakl_api_key = get_portal_setting('mirakl_api_key')
    marketplace_export_defaults = _get_marketplace_export_defaults()
    return render_template(
        'marketplace_export.html',
        mirakl_base_url=mirakl_base_url,
        mirakl_shop_id=mirakl_shop_id,
        mirakl_api_key_set=bool(mirakl_api_key),
        marketplace_export_defaults=marketplace_export_defaults,
    )


@marketplace_bp.route('/export-defaults', methods=['GET'])
def marketplace_export_defaults_get():
    return jsonify({'success': True, 'defaults': _get_marketplace_export_defaults()}), 200


@marketplace_bp.route('/export-defaults', methods=['POST'])
def marketplace_export_defaults_save():
    data = request.get_json() or {}
    defaults = _normalize_marketplace_export_defaults(data.get('defaults'))

    db = get_db()
    cursor = None
    try:
        cursor = db.cursor()
        _upsert_portal_setting(cursor, 'marketplace_export_default_description_mode', defaults['description_mode'])
        _upsert_portal_setting(cursor, 'marketplace_export_default_mkp_eccn', defaults['mkp_eccn'] or '')
        _upsert_portal_setting(cursor, 'marketplace_export_default_mkp_product_unit', defaults['mkp_product_unit'] or '')
        _upsert_portal_setting(cursor, 'marketplace_export_default_mkp_package_content', str(defaults['mkp_package_content']))
        _upsert_portal_setting(cursor, 'marketplace_export_default_mkp_package_content_unit', defaults['mkp_package_content_unit'] or '')
        _upsert_portal_setting(cursor, 'marketplace_export_default_mkp_third_level', defaults['mkp_third_level'] or '')
        _upsert_portal_setting(cursor, 'marketplace_export_default_mkp_dangerous', '1' if defaults['mkp_dangerous'] else '0')
        _upsert_portal_setting(cursor, 'marketplace_export_default_mkp_serialized', '1' if defaults['mkp_serialized'] else '0')
        _upsert_portal_setting(cursor, 'marketplace_export_default_mkp_log_card', '1' if defaults['mkp_log_card'] else '0')
        _upsert_portal_setting(cursor, 'marketplace_export_default_mkp_easaf1', '1' if defaults['mkp_easaf1'] else '0')
        db.commit()
    except Exception as exc:
        logger.exception("Failed to save marketplace export defaults")
        db.rollback()
        return jsonify({'success': False, 'error': str(exc)}), 500
    finally:
        if cursor:
            cursor.close()

    return jsonify({'success': True, 'defaults': defaults}), 200


@marketplace_bp.route('/mirakl', methods=['GET'])
def mirakl_connection_page():
    """Render the Mirakl connection page."""
    portal_base_url = get_portal_setting('mirakl_base_url')
    portal_shop_id = get_portal_setting('mirakl_shop_id')
    portal_api_key = get_portal_setting('mirakl_api_key')
    env_base_url = os.getenv('MIRAKL_BASE_URL')
    env_shop_id = os.getenv('MIRAKL_SHOP_ID')
    env_api_key = os.getenv('MIRAKL_API_KEY')
    display_base_url = portal_base_url or env_base_url
    display_shop_id = portal_shop_id or env_shop_id
    return render_template(
        'marketplace_mirakl.html',
        mirakl_base_url=display_base_url,
        mirakl_shop_id=display_shop_id,
        mirakl_api_key_set=bool(portal_api_key or env_api_key),
        mirakl_env_configured=bool(env_base_url or env_shop_id or env_api_key),
    )


@marketplace_bp.route('/mirakl/p31', methods=['GET'])
def mirakl_p31_page():
    return render_template('marketplace_mirakl_p31.html')


@marketplace_bp.route('/mirakl/p31/lookup', methods=['POST'])
def mirakl_p31_lookup():
    client, error = _get_mirakl_client()
    if error:
        return jsonify({'success': False, 'error': error}), 400

    data = request.get_json() or {}
    reference_type = _coerce_text(data.get('reference_type'), default='mpnTitle')
    references = list(dict.fromkeys(_parse_reference_input(data.get('references'))))

    if not references:
        return jsonify({'success': False, 'error': 'At least one reference is required.'}), 400
    if len(references) > 500:
        return jsonify({'success': False, 'error': 'Please limit lookups to 500 references at a time.'}), 400

    crm_rows = _fetch_marketplace_parts_by_references(references)
    crm_lookup = {row['requested_reference']: row.get('crm_part') for row in crm_rows}
    matched_products = {}
    candidate_map = {ref: [] for ref in references}

    try:
        for start in range(0, len(references), 100):
            batch = references[start:start + 100]
            response = client.get_products_by_references([f'{reference_type}|{value}' for value in batch])
            products = _extract_mirakl_product_rows(response)
            matched_products.update(_match_mirakl_products_by_reference(products, batch))
            for ref in batch:
                candidate_map[ref].extend(_collect_matching_candidates(products, ref))
    except MiraklError as exc:
        logger.exception("Mirakl P31 lookup failed")
        return jsonify({'success': False, 'error': str(exc)}), 502

    rows = []
    for ref in references:
        crm_part = crm_lookup.get(ref)
        product = matched_products.get(ref)
        product_sku = _extract_mirakl_product_sku(product or {})
        rows.append({
            'requested_reference': ref,
            'reference_type': reference_type,
            'crm_part': crm_part,
            'product_found': bool(product),
            'resolved_product_id': product_sku,
            'resolved_product_id_type': 'SKU' if product_sku else '',
            'resolved_product_label': _extract_mirakl_product_label(product or {}),
            'mirakl_product': product or {},
            'mirakl_candidates': candidate_map.get(ref, [])[:5],
        })

    return jsonify({
        'success': True,
        'reference_type': reference_type,
        'rows': rows,
        'summary': {
            'requested': len(references),
            'crm_matched': sum(1 for row in rows if row.get('crm_part')),
            'product_found': sum(1 for row in rows if row.get('product_found')),
            'offer_ready': sum(1 for row in rows if row.get('crm_part') and row.get('resolved_product_id')),
        },
    }), 200


@marketplace_bp.route('/categories', methods=['GET'])
def get_marketplace_categories():
    """Get all available Airbus Marketplace categories"""
    try:
        categories = get_available_categories()
        return jsonify({'categories': categories}), 200
    except Exception as e:
        logger.exception("Error getting marketplace categories")
        return jsonify({'error': str(e)}), 500


@marketplace_bp.route('/categorization-tool', methods=['GET'])
def categorization_tool_page():
    """Render the marketplace categorization tool page"""
    return render_template('marketplace_category_tool.html')


@marketplace_bp.route('/categorization-tool/suggest', methods=['POST'])
def categorization_tool_suggest():
    data = request.get_json() or {}
    part_number = (data.get('part_number') or '').strip()
    description = (data.get('description') or '').strip()
    use_perplexity = bool(data.get('use_perplexity', True))

    if not part_number:
        return jsonify({'success': False, 'error': 'part_number is required'}), 400

    categories = get_available_categories()
    category_list = "\n".join([f"- {cat}" for cat in categories])

    if use_perplexity:
        client = _get_perplexity_client()
        if not client:
            return jsonify({'success': False, 'error': 'Perplexity API key not configured'}), 400

        system_message = f"""You are an aviation hardware analyst. Return ONLY valid JSON.

Choose one category from this exact list:
{category_list}

Rules:
1. Output ONLY JSON, no markdown
2. If unsure, use "Marketplace Categories/Hardware and Electrical/Miscellaneous"
3. Reasoning must be max 12 words, no quotes
4. Suggest 0-3 prefixes to apply for mass updates, or an empty list

Output format:
{{
  "category": "Marketplace Categories/Hardware and Electrical/CategoryName",
  "confidence": "high|medium|low",
  "reasoning": "short reason",
  "prefixes": ["EXAMPLE", "SERIES"]
}}"""

        user_message = f"""Part number: {part_number}
Description: {description or "None"}

Suggest a category and any useful prefixes or series identifiers."""

        try:
            response = client.chat.completions.create(
                model="sonar-pro",
                messages=[
                    {"role": "system", "content": system_message},
                    {"role": "user", "content": user_message},
                ],
                temperature=0.2,
                max_tokens=500,
            )
            raw_content = response.choices[0].message.content.strip()
            try:
                result = json.loads(raw_content)
            except json.JSONDecodeError:
                extracted = _extract_json_object(raw_content)
                result = json.loads(extracted) if extracted else {}
        except Exception as exc:
            logger.exception("Perplexity suggestion failed")
            return jsonify({'success': False, 'error': str(exc)}), 500
    else:
        result = suggest_marketplace_category(part_number, description, '')

    category = result.get('category') if isinstance(result, dict) else None
    confidence = result.get('confidence', 'low') if isinstance(result, dict) else 'low'
    reasoning = result.get('reasoning', '') if isinstance(result, dict) else ''
    prefixes = _normalize_prefixes(result.get('prefixes') if isinstance(result, dict) else [])

    if category not in categories:
        category = "Marketplace Categories/Hardware and Electrical/Miscellaneous"
        confidence = "low"
        if not reasoning:
            reasoning = "invalid_category"

    if not prefixes and not use_perplexity:
        prefixes = _guess_prefixes_from_part_number(part_number)

    return jsonify({
        'success': True,
        'category': category,
        'confidence': confidence,
        'reasoning': reasoning,
        'prefixes': prefixes,
    }), 200


@marketplace_bp.route('/categorization-tool/uncategorized', methods=['GET'])
def categorization_tool_uncategorized():
    limit = request.args.get('limit', 50)
    try:
        limit = max(1, min(int(limit), 200))
    except (TypeError, ValueError):
        limit = 50

    db = get_db()
    cursor = db.cursor()
    cursor.execute(
        """
        SELECT base_part_number, part_number
        FROM part_numbers
        WHERE mkp_category IS NULL OR mkp_category = ''
        ORDER BY part_number
        LIMIT ?
        """,
        (limit,)
    )
    rows = cursor.fetchall()
    parts = [
        {
            'base_part_number': row['base_part_number'],
            'part_number': row['part_number'] or row['base_part_number'],
        }
        for row in rows
    ]
    return jsonify({'success': True, 'parts': parts}), 200


@marketplace_bp.route('/categorization-tool/uncategorized-search', methods=['GET'])
def categorization_tool_uncategorized_search():
    query = (request.args.get('q') or '').strip().lower()
    limit = request.args.get('limit', 200)
    try:
        limit = max(1, min(int(limit), 200))
    except (TypeError, ValueError):
        limit = 200

    if not query:
        return jsonify({'success': True, 'parts': []}), 200

    like_value = f"%{query}%"
    db = get_db()
    cursor = db.cursor()
    cursor.execute(
        """
        SELECT base_part_number, part_number
        FROM part_numbers
        WHERE (mkp_category IS NULL OR mkp_category = '')
          AND (LOWER(part_number) LIKE ? OR LOWER(base_part_number) LIKE ?)
        ORDER BY part_number
        LIMIT ?
        """,
        (like_value, like_value, limit)
    )
    rows = cursor.fetchall()
    parts = [
        {
            'base_part_number': row['base_part_number'],
            'part_number': row['part_number'] or row['base_part_number'],
        }
        for row in rows
    ]
    return jsonify({'success': True, 'parts': parts}), 200


@marketplace_bp.route('/categorization-tool/uncategorized-ranked', methods=['GET'])
def categorization_tool_uncategorized_ranked():
    source = (request.args.get('source') or 'stock_qty').strip().lower()
    limit = request.args.get('limit', 100)
    try:
        limit = max(1, min(int(limit), 500))
    except (TypeError, ValueError):
        limit = 100

    supported_sources = {
        'stock_qty': {
            'title': 'uncategorized parts ranked by stock quantity',
            'filter_sql': 'COALESCE(stock.stock_quantity, 0) > 0',
            'order_sql': 'COALESCE(stock.stock_quantity, 0) DESC, COALESCE(stock.stock_value, 0) DESC, pn.part_number ASC',
        },
        'stock_value': {
            'title': 'uncategorized parts ranked by stock value',
            'filter_sql': 'COALESCE(stock.stock_value, 0) > 0',
            'order_sql': 'COALESCE(stock.stock_value, 0) DESC, COALESCE(stock.stock_quantity, 0) DESC, pn.part_number ASC',
        },
        'sales_frequency': {
            'title': 'uncategorized parts ranked by sales frequency',
            'filter_sql': 'COALESCE(sales.sales_order_count, 0) > 0',
            'order_sql': 'COALESCE(sales.sales_order_count, 0) DESC, COALESCE(sales.quantity_sold, 0) DESC, pn.part_number ASC',
        },
    }
    config = supported_sources.get(source)
    if not config:
        return jsonify({'success': False, 'error': f'Unsupported source: {source}'}), 400

    db = get_db()
    cursor = db.cursor()
    cursor.execute(
        f"""
        WITH stock AS (
            SELECT
                sm.base_part_number,
                COALESCE(SUM(sm.available_quantity), 0) AS stock_quantity,
                COALESCE(SUM(sm.available_quantity * COALESCE(sm.cost_per_unit, 0)), 0) AS stock_value
            FROM stock_movements sm
            WHERE sm.movement_type = 'IN'
              AND sm.available_quantity > 0
            GROUP BY sm.base_part_number
        ),
        sales AS (
            SELECT
                sol.base_part_number,
                COUNT(*) AS sales_order_count,
                COALESCE(SUM(sol.quantity), 0) AS quantity_sold,
                MAX(so.date_entered) AS last_sale_date
            FROM sales_order_lines sol
            JOIN sales_orders so ON so.id = sol.sales_order_id
            GROUP BY sol.base_part_number
        )
        SELECT
            pn.base_part_number,
            COALESCE(pn.part_number, pn.base_part_number) AS part_number,
            COALESCE(stock.stock_quantity, 0) AS stock_quantity,
            COALESCE(stock.stock_value, 0) AS stock_value,
            COALESCE(sales.sales_order_count, 0) AS sales_order_count,
            COALESCE(sales.quantity_sold, 0) AS quantity_sold,
            sales.last_sale_date
        FROM part_numbers pn
        LEFT JOIN stock ON stock.base_part_number = pn.base_part_number
        LEFT JOIN sales ON sales.base_part_number = pn.base_part_number
        WHERE (pn.mkp_category IS NULL OR pn.mkp_category = '')
          AND {config['filter_sql']}
        ORDER BY {config['order_sql']}
        LIMIT ?
        """,
        (limit,),
    )
    rows = cursor.fetchall() or []
    parts = [
        {
            'base_part_number': row['base_part_number'],
            'part_number': row['part_number'] or row['base_part_number'],
            'stock_quantity': float(row['stock_quantity'] or 0),
            'stock_value': float(row['stock_value'] or 0),
            'sales_order_count': int(row['sales_order_count'] or 0),
            'quantity_sold': float(row['quantity_sold'] or 0),
            'last_sale_date': row['last_sale_date'].isoformat() if row['last_sale_date'] else None,
        }
        for row in rows
    ]
    return jsonify({
        'success': True,
        'source': source,
        'title': config['title'],
        'parts': parts,
    }), 200


@marketplace_bp.route('/categorization-tool/prefix-preview', methods=['POST'])
def categorization_tool_prefix_preview():
    data = request.get_json() or {}
    prefixes = _normalize_prefixes(data.get('prefixes', []))
    only_uncategorized = bool(data.get('only_uncategorized', True))

    if not prefixes:
        return jsonify({'success': True, 'total': 0, 'sample': []}), 200

    conditions, params = _build_prefix_conditions(prefixes)
    where_clause = " OR ".join(conditions)
    filters = [f"({where_clause})"]
    if only_uncategorized:
        filters.append("(pn.mkp_category IS NULL OR pn.mkp_category = '')")

    where_sql = " AND ".join(filters)

    db = get_db()
    cursor = db.cursor()

    count_query = f"SELECT COUNT(*) AS total FROM part_numbers pn WHERE {where_sql}"
    cursor.execute(count_query, params)
    total_row = cursor.fetchone()
    total = total_row['total'] if total_row else 0

    sample_query = f"""
        SELECT pn.base_part_number, pn.part_number, pn.mkp_category
        FROM part_numbers pn
        WHERE {where_sql}
        ORDER BY pn.part_number
        LIMIT 20
    """
    cursor.execute(sample_query, params)
    rows = cursor.fetchall()
    sample = [
        {
            'base_part_number': row['base_part_number'],
            'part_number': row['part_number'],
            'mkp_category': row['mkp_category'],
        }
        for row in rows
    ]

    return jsonify({'success': True, 'total': total, 'sample': sample}), 200


@marketplace_bp.route('/categorization-tool/apply-prefix', methods=['POST'])
def categorization_tool_apply_prefix():
    data = request.get_json() or {}
    prefixes = _normalize_prefixes(data.get('prefixes', []))
    category = (data.get('category') or '').strip()
    only_uncategorized = bool(data.get('only_uncategorized', True))

    if not prefixes:
        return jsonify({'success': False, 'error': 'prefixes are required'}), 400
    if not category:
        return jsonify({'success': False, 'error': 'category is required'}), 400

    valid_categories = get_available_categories()
    if category not in valid_categories:
        return jsonify({'success': False, 'error': 'Invalid category'}), 400

    conditions, params = _build_prefix_conditions(prefixes)
    where_clause = " OR ".join(conditions)
    filters = [f"({where_clause})"]
    if only_uncategorized:
        filters.append("(mkp_category IS NULL OR mkp_category = '')")
    where_sql = " AND ".join(filters)

    db = get_db()
    cursor = db.cursor()
    update_query = f"UPDATE part_numbers pn SET mkp_category = ? WHERE {where_sql}"
    cursor.execute(update_query, [category] + params)
    db.commit()

    return jsonify({'success': True, 'updated_count': cursor.rowcount}), 200


@marketplace_bp.route('/categorization-tool/apply-list', methods=['POST'])
def categorization_tool_apply_list():
    data = request.get_json() or {}
    base_part_numbers = data.get('base_part_numbers') or []
    category = (data.get('category') or '').strip()

    if not base_part_numbers:
        return jsonify({'success': False, 'error': 'base_part_numbers are required'}), 400
    if not category:
        return jsonify({'success': False, 'error': 'category is required'}), 400

    valid_categories = get_available_categories()
    if category not in valid_categories:
        return jsonify({'success': False, 'error': 'Invalid category'}), 400

    placeholders = ','.join('?' * len(base_part_numbers))
    query = f"UPDATE part_numbers SET mkp_category = ? WHERE base_part_number IN ({placeholders})"

    db = get_db()
    cursor = db.cursor()
    cursor.execute(query, [category] + base_part_numbers)
    db.commit()

    return jsonify({'success': True, 'updated_count': cursor.rowcount}), 200


@marketplace_bp.route('/suggest-category', methods=['POST'])
def suggest_category():
    """
    Suggest category for a single part

    POST body:
    {
        "part_number": "ABC123",
        "description": "Optional description",
        "additional_info": "Optional additional context"
    }
    """
    try:
        data = request.get_json()
        part_number = data.get('part_number')
        description = data.get('description', '')
        additional_info = data.get('additional_info', '')

        if not part_number:
            return jsonify({'error': 'part_number is required'}), 400

        suggestion = suggest_marketplace_category(
            part_number,
            description,
            additional_info
        )

        if not suggestion:
            return jsonify({'error': 'Unable to suggest category'}), 500

        return jsonify(suggestion), 200

    except Exception as e:
        logger.exception("Error suggesting category")
        return jsonify({'error': str(e)}), 500


@marketplace_bp.route('/suggest-categories-batch', methods=['POST'])
def suggest_categories_batch_route():
    """
    Suggest categories for multiple parts

    POST body:
    {
        "parts": [
            {"part_number": "ABC123", "description": "..."},
            {"part_number": "XYZ789", "description": "..."}
        ]
    }
    """
    try:
        data = request.get_json()
        parts = data.get('parts', [])

        if not parts:
            return jsonify({'error': 'parts array is required'}), 400

        results = suggest_categories_batch(parts)

        return jsonify({'results': results}), 200

    except Exception as e:
        logger.exception("Error suggesting categories in batch")
        return jsonify({'error': str(e)}), 500


@marketplace_bp.route('/update-category/<base_part_number>', methods=['POST'])
def update_part_category(base_part_number):
    """
    Update marketplace category for a part

    POST body:
    {
        "mkp_category": "Marketplace Categories/Hardware and Electrical/Bolts"
    }
    """
    try:
        data = request.get_json()
        mkp_category = data.get('mkp_category')

        if not mkp_category:
            return jsonify({'error': 'mkp_category is required'}), 400

        # Validate category is in allowed list
        valid_categories = get_available_categories()
        if mkp_category not in valid_categories:
            return jsonify({'error': 'Invalid category'}), 400

        # Update part in database
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            "UPDATE part_numbers SET mkp_category = ? WHERE base_part_number = ?",
            (mkp_category, base_part_number)
        )
        db.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': 'Part not found'}), 404

        return jsonify({'success': True, 'message': 'Category updated'}), 200

    except Exception as e:
        logger.exception("Error updating part category")
        if db:
            db.rollback()
        return jsonify({'error': str(e)}), 500


@marketplace_bp.route('/update-marketplace-fields/<base_part_number>', methods=['POST'])
def update_marketplace_fields(base_part_number):
    """
    Update all marketplace fields for a part

    POST body:
    {
        "mkp_category": "Marketplace Categories/Hardware and Electrical/Bolts",
        "mkp_description": "Custom description",
        "mkp_name": "Custom name",
        "mkp_product_summary": "Product summary text",
        "mkp_product_presentation": "Product presentation text",
        "mkp_product_unit": "EA",
        "mkp_package_content": 1,
        "mkp_package_content_unit": "EA",
        "mkp_third_level": "EA",
        "mkp_dangerous": false,
        "mkp_eccn": "",
        "mkp_serialized": false,
        "mkp_log_card": false,
        "mkp_easaf1": false
    }
    """
    db = None
    try:
        data = request.get_json() or {}

        # Validate category if provided
        mkp_category = data.get('mkp_category')
        if mkp_category:
            valid_categories = get_available_categories()
            if mkp_category not in valid_categories:
                return jsonify({'error': 'Invalid category'}), 400

        # Build update query dynamically based on provided fields
        allowed_fields = [
            'mkp_category', 'mkp_description', 'mkp_name', 'mkp_product_summary',
            'mkp_product_presentation', 'mkp_product_unit', 'mkp_package_content',
            'mkp_package_content_unit', 'mkp_third_level', 'mkp_dangerous',
            'mkp_eccn', 'mkp_serialized', 'mkp_log_card', 'mkp_easaf1'
        ]

        updates = []
        params = []
        for field in allowed_fields:
            if field in data:
                updates.append(f"{field} = ?")
                value = data[field]
                if field in ('mkp_dangerous', 'mkp_serialized', 'mkp_log_card', 'mkp_easaf1'):
                    if value is not None:
                        parsed_bool = _parse_import_bool(value)
                        if parsed_bool is None:
                            return jsonify({'error': f'Invalid boolean value for {field}'}), 400
                        value = parsed_bool
                params.append(value)

        if not updates:
            return jsonify({'error': 'No fields to update'}), 400

        params.append(base_part_number)
        query = f"UPDATE part_numbers SET {', '.join(updates)} WHERE base_part_number = ?"

        db = get_db()
        cursor = db.cursor()
        cursor.execute(query, params)
        db.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': 'Part not found'}), 404

        return jsonify({'success': True, 'message': 'Marketplace fields updated'}), 200

    except Exception as e:
        logger.exception("Error updating marketplace fields")
        if db:
            db.rollback()
        return jsonify({'error': str(e)}), 500


@marketplace_bp.route('/get-parts-for-export', methods=['POST'])
def get_parts_for_export():
    """
    Get filtered parts with portal pricing data for export preview

    POST body:
    {
        "stock_filter": "all|stock_only|no_stock",
        "category_filter": "all|categorized_only|missing_only",
        "pricing_only": true/false,
        "part_number_search": optional part number search
    }
    """
    try:
        data = request.get_json() or {}

        stock_filter = data.get('stock_filter', 'all')
        category_filter = data.get('category_filter', 'all')
        pricing_only = data.get('pricing_only', False)
        part_number_search = data.get('part_number_search', '').strip()
        include_non_hqpl_alts = _coerce_bool(data.get('include_non_hqpl_alts'), default=False)
        include_alt_stock_rollup = _coerce_bool(data.get('include_alt_stock_rollup'), default=False)
        source_mode = _coerce_text(data.get('source_mode'), default='filters')
        if source_mode not in ('filters', 'baseline'):
            source_mode = 'filters'
        selected_base_part_numbers = [
            str(value).strip()
            for value in (data.get('selected_base_part_numbers') or [])
            if str(value).strip()
        ]
        max_results = _coerce_int(data.get('max_results'), default=0)
        if max_results < 0:
            max_results = 0

        logger.info(
            "Marketplace export parts request: source_mode=%s stock_filter=%s category_filter=%s "
            "pricing_only=%s part_number_search=%s selected_count=%s max_results=%s",
            source_mode,
            stock_filter,
            category_filter,
            pricing_only,
            part_number_search or "<none>",
            len(selected_base_part_numbers),
            max_results or "<none>",
        )

        customer_id = _get_marketplace_customer_id()

        db = get_db()
        cursor = db.cursor()

        query = """
            SELECT
                pn.base_part_number,
                pn.part_number,
                marketplace_mfg.manufacturer_name,
                pn.mkp_category,
                pn.mkp_description,
                pn.mkp_name,
                pn.mkp_product_summary,
                pn.mkp_product_presentation,
                pn.mkp_product_unit,
                pn.mkp_package_content,
                pn.mkp_package_content_unit,
                pn.mkp_third_level,
                pn.mkp_dangerous,
                pn.mkp_eccn,
                pn.mkp_serialized,
                pn.mkp_log_card,
                pn.mkp_easaf1,
                pn.mkp_offer_product_id,
                pn.mkp_offer_product_id_type
            FROM part_numbers pn
            """ + _MARKETPLACE_MANUFACTURER_JOIN + """
            WHERE 1=1
        """

        params = [AIRBUS_ROTARY_APPROVAL_LIST_TYPE]
        query += " AND " + _build_rotary_hqpl_exists_clause('pn')

        if selected_base_part_numbers:
            placeholders = ','.join('?' * len(selected_base_part_numbers))
            query += f" AND pn.base_part_number IN ({placeholders})"
            params.extend(selected_base_part_numbers)

        # Apply category filter
        if category_filter == 'categorized_only':
            query += " AND pn.mkp_category IS NOT NULL AND pn.mkp_category != ''"
        elif category_filter == 'missing_only':
            query += " AND (pn.mkp_category IS NULL OR pn.mkp_category = '')"

        # Apply pricing filter (mirror portal recency rules to avoid loading all parts)
        if pricing_only:
            so_months = int(get_portal_setting('sales_order_recency_months', 6))
            vq_months = int(get_portal_setting('vq_recency_months', 12))
            po_months = int(get_portal_setting('po_recency_months', 12))
            cq_months = int(get_portal_setting('cq_recency_months', 6))
            min_stock = int(get_portal_setting('min_stock_threshold', 1))
            show_estimates = bool(int(get_portal_setting('show_estimated_prices', 1)))

            today = datetime.utcnow().date()
            so_cutoff = _months_ago(today, so_months)
            vq_cutoff = _months_ago(today, vq_months)
            po_cutoff = _months_ago(today, po_months)
            cq_cutoff = _months_ago(today, cq_months)

            pricing_filters = [
                """
                EXISTS (
                    SELECT 1
                    FROM stock_movements sm
                    WHERE sm.base_part_number = pn.base_part_number
                      AND sm.movement_type = 'IN'
                      AND sm.available_quantity >= ?
                )
                """,
                """
                EXISTS (
                    SELECT 1
                    FROM cq_lines cl
                    JOIN cqs c ON cl.cq_id = c.id
                    WHERE cl.base_part_number = pn.base_part_number
                      AND c.entry_date >= ?
                      AND cl.unit_price > 0
                      AND cl.is_no_quote = FALSE
                )
                """,
                """
                EXISTS (
                    SELECT 1
                    FROM sales_order_lines sol
                    JOIN sales_orders so ON sol.sales_order_id = so.id
                    WHERE sol.base_part_number = pn.base_part_number
                      AND so.date_entered >= ?
                      AND sol.price > 0
                )
                """,
                """
                EXISTS (
                    SELECT 1
                    FROM customer_quote_lines cql
                    JOIN parts_list_lines pll ON pll.id = cql.parts_list_line_id
                    WHERE pll.base_part_number = pn.base_part_number
                      AND cql.quoted_status = 'quoted'
                      AND cql.quote_price_gbp > 0
                      AND COALESCE(cql.is_no_bid, 0) = 0
                      AND cql.date_created >= ?
                )
                """,
                """
                EXISTS (
                    SELECT 1
                    FROM parts_list_supplier_quote_lines sql
                    JOIN parts_list_supplier_quotes sq ON sq.id = sql.supplier_quote_id
                    JOIN parts_list_lines pll ON pll.id = sql.parts_list_line_id
                    WHERE pll.base_part_number = pn.base_part_number
                      AND sql.unit_price > 0
                      AND (sql.is_no_bid = FALSE OR sql.is_no_bid IS NULL)
                      AND COALESCE(sq.quote_date, sq.date_created) >= ?
                )
                """,
            ]

            params.extend([
                min_stock,
                cq_cutoff,
                so_cutoff,
                cq_cutoff,
                vq_cutoff,
            ])

            if customer_id:
                pricing_filters.append(
                    """
                    EXISTS (
                        SELECT 1
                        FROM portal_customer_pricing pcp
                        WHERE pcp.customer_id = ?
                          AND pcp.base_part_number = pn.base_part_number
                          AND pcp.is_active = TRUE
                          AND (pcp.valid_from IS NULL OR pcp.valid_from <= ?)
                          AND (pcp.valid_until IS NULL OR pcp.valid_until >= ?)
                    )
                    """
                )
                params.extend([customer_id, today, today])

            if show_estimates:
                pricing_filters.extend([
                    """
                    EXISTS (
                        SELECT 1
                        FROM vq_lines vl
                        JOIN vqs v ON vl.vq_id = v.id
                        WHERE vl.base_part_number = pn.base_part_number
                          AND v.entry_date >= ?
                          AND vl.vendor_price > 0
                    )
                    """,
                    """
                    EXISTS (
                        SELECT 1
                        FROM purchase_order_lines pol
                        JOIN purchase_orders po ON pol.purchase_order_id = po.id
                        WHERE pol.base_part_number = pn.base_part_number
                          AND po.date_issued >= ?
                          AND pol.price > 0
                    )
                    """,
                ])
                params.extend([vq_cutoff, po_cutoff])

            query += " AND (" + " OR ".join(pricing_filters) + ")"

        # Apply part number search
        if part_number_search:
            query += " AND (pn.part_number LIKE ? OR pn.base_part_number LIKE ?)"
            params.append(f"%{part_number_search}%")
            params.append(f"%{part_number_search}%")

        query += " ORDER BY pn.part_number"
        if max_results:
            query += " LIMIT ?"
            params.append(max_results)

        cursor.execute(query, params)
        rows = cursor.fetchall()
        logger.info("Marketplace export query returned %s parts", len(rows))

        rotary_alt_entries_map = _get_global_alternative_entries_map(
            cursor,
            [row['base_part_number'] for row in rows],
            rotary_only=True,
        )
        all_alt_entries_map = _get_global_alternative_entries_map(
            cursor,
            [row['base_part_number'] for row in rows],
            rotary_only=False,
        )
        rotary_alt_map = {
            source_base: [entry['part_number'] for entry in entries]
            for source_base, entries in rotary_alt_entries_map.items()
        }
        all_alt_map = {
            source_base: [entry['part_number'] for entry in entries]
            for source_base, entries in all_alt_entries_map.items()
        }

        parts_payload = [
            {'part_number': row['part_number'] or row['base_part_number'], 'quantity': 1}
            for row in rows
        ]
        estimates_start = time.monotonic()
        estimates = _get_portal_estimates(parts_payload, customer_id)
        logger.info(
            "Marketplace export estimates: requested=%s returned=%s took=%.2fs",
            len(parts_payload),
            len(estimates),
            time.monotonic() - estimates_start,
        )
        rotary_alt_stock_rollup_map = _get_alt_stock_rollup_map(customer_id, rotary_alt_entries_map)
        all_alt_stock_rollup_map = _get_alt_stock_rollup_map(customer_id, all_alt_entries_map)

        parts = []
        filtered_stock = 0
        filtered_no_stock = 0
        filtered_pricing = 0
        reference_summary = {
            'exact': 0,
            'exact_normalized': 0,
            'alias': 0,
            'alias_normalized': 0,
            'unknown': 0,
            'missing': 0,
        }
        for row in rows:
            base_part_number = row['base_part_number']
            part_number = row['part_number'] or base_part_number
            estimate = estimates.get(base_part_number) or estimates.get(row['part_number'])
            selected_alt_rollup = _select_alt_stock_rollup(
                include_non_hqpl_alts,
                rotary_alt_stock_rollup_map.get(base_part_number),
                all_alt_stock_rollup_map.get(base_part_number),
            )
            effective_estimate = (
                _apply_alt_stock_rollup_to_estimate(estimate, selected_alt_rollup)
                if include_alt_stock_rollup
                else (estimate or {})
            )
            estimated_price = effective_estimate.get('estimated_price')
            in_stock = bool(effective_estimate.get('in_stock')) if effective_estimate else False
            stock_qty = effective_estimate.get('stock_quantity') if effective_estimate else None

            if stock_filter == 'stock_only' and not in_stock:
                filtered_stock += 1
                continue
            if stock_filter == 'no_stock' and in_stock:
                filtered_no_stock += 1
                continue
            if pricing_only and estimated_price is None:
                filtered_pricing += 1
                continue

            reference_resolution = _resolve_airbus_mpn_title(part_number)
            reference_summary[reference_resolution['reference_status']] = (
                reference_summary.get(reference_resolution['reference_status'], 0) + 1
            )

            parts.append({
                'base_part_number': base_part_number,
                'part_number': row['part_number'],
                'mkp_category': row['mkp_category'],
                'mkp_description': row['mkp_description'],
                'mkp_name': row['mkp_name'],
                'mkp_product_summary': row['mkp_product_summary'],
                'mkp_product_presentation': row['mkp_product_presentation'],
                'mkp_product_unit': row['mkp_product_unit'],
                'mkp_package_content': row['mkp_package_content'],
                'mkp_package_content_unit': row['mkp_package_content_unit'],
                'mkp_third_level': row['mkp_third_level'],
                'mkp_dangerous': row['mkp_dangerous'],
                'mkp_eccn': row['mkp_eccn'],
                'mkp_serialized': row['mkp_serialized'],
                'mkp_log_card': row['mkp_log_card'],
                'mkp_easaf1': row['mkp_easaf1'],
                'mkp_offer_product_id': row['mkp_offer_product_id'],
                'mkp_offer_product_id_type': row['mkp_offer_product_id_type'],
                'all_global_alt_part_numbers': all_alt_map.get(base_part_number, []),
                'rotary_hqpl_alt_part_numbers': rotary_alt_map.get(base_part_number, []),
                'all_alt_stock_rollup': all_alt_stock_rollup_map.get(base_part_number, {}),
                'rotary_alt_stock_rollup': rotary_alt_stock_rollup_map.get(base_part_number, {}),
                'description': '',
                'manufacturer': _coerce_text(row['manufacturer_name'], default=''),
                'estimated_price': estimate.get('estimated_price') if estimate else None,
                'price_source': estimate.get('price_source') if estimate else None,
                'estimated_lead_days': estimate.get('estimated_lead_days') if estimate else None,
                'currency': estimate.get('currency') if estimate else None,
                'in_stock': in_stock,
                'stock_qty': stock_qty if stock_qty is not None else 0,
                'resolved_mpn_title': reference_resolution['resolved_mpn_title'],
                'reference_status': reference_resolution['reference_status'],
                'reference_source': reference_resolution['reference_source'],
            })

        logger.info(
            "Marketplace export parts response: total=%s returned=%s "
            "filtered_stock=%s filtered_no_stock=%s filtered_pricing=%s",
            len(rows),
            len(parts),
            filtered_stock,
            filtered_no_stock,
            filtered_pricing,
        )
        return jsonify({
            'success': True,
            'parts': parts,
            'reference_summary': reference_summary,
        }), 200

    except Exception as e:
        logger.exception("Error getting parts for export")
        return jsonify({'success': False, 'error': str(e)}), 500

@marketplace_bp.route('/export', methods=['POST'])
def export_to_marketplace():
    """
    Export selected parts to Airbus Marketplace format with pricing.

    Accepts either JSON request bodies or legacy form posts containing an
    ``export_data`` field so large exports do not hit form parser limits.
    """
    try:
        export_data = request.get_json(silent=True)
        if export_data is None:
            export_data_str = request.form.get('export_data')
            if not export_data_str:
                return jsonify({'error': 'No export data provided'}), 400
            export_data = json.loads(export_data_str)
        base_part_numbers = export_data.get('base_part_numbers', [])
        default_quantity = int(export_data.get('default_quantity') or 1)
        skip_invalid_mandatory = _coerce_bool(export_data.get('skip_invalid_mandatory'), default=False)
        include_non_hqpl_alts = _coerce_bool(export_data.get('include_non_hqpl_alts'), default=False)
        include_alt_stock_rollup = _coerce_bool(export_data.get('include_alt_stock_rollup'), default=False)
        export_mode = _coerce_text(export_data.get('export_mode'), default='products')
        debug_offer_export = _coerce_bool(export_data.get('debug_offer_export'), default=False)
        if export_mode not in ('products', 'offers'):
            export_mode = 'products'
        export_defaults = _normalize_marketplace_export_defaults(export_data.get('defaults'))
        source_mode = _coerce_text(export_data.get('source_mode'), default='filters')
        if source_mode not in ('filters', 'baseline'):
            source_mode = 'filters'
        baseline_rows = export_data.get('baseline_rows') or {}

        if not base_part_numbers:
            return jsonify({'error': 'No parts selected for export'}), 400

        db = get_db()
        cursor = db.cursor()

        placeholders = ','.join('?' * len(base_part_numbers))
        query = f"""
            SELECT
                pn.base_part_number,
                pn.part_number,
                marketplace_mfg.manufacturer_name,
                pn.mkp_category,
                pn.mkp_description,
                pn.mkp_name,
                pn.mkp_product_summary,
                pn.mkp_product_presentation,
                pn.mkp_product_unit,
                pn.mkp_package_content,
                pn.mkp_package_content_unit,
                pn.mkp_third_level,
                pn.mkp_dangerous,
                pn.mkp_eccn,
                pn.mkp_serialized,
                pn.mkp_log_card,
                pn.mkp_easaf1,
                pn.mkp_offer_product_id,
                pn.mkp_offer_product_id_type
            FROM part_numbers pn
            {_MARKETPLACE_MANUFACTURER_JOIN}
            WHERE pn.base_part_number IN ({placeholders})
              AND {_build_rotary_hqpl_exists_clause('pn')}
        """

        cursor.execute(query, base_part_numbers + [AIRBUS_ROTARY_APPROVAL_LIST_TYPE])
        rows = cursor.fetchall()

        if not rows:
            return jsonify({'error': 'No rotary HQPL-approved parts found to export'}), 404

        rotary_alt_entries_map = _get_global_alternative_entries_map(
            cursor,
            [row['base_part_number'] for row in rows],
            rotary_only=True,
        )
        all_alt_entries_map = _get_global_alternative_entries_map(
            cursor,
            [row['base_part_number'] for row in rows],
            rotary_only=False,
        )
        rotary_alt_map = {
            source_base: [entry['part_number'] for entry in entries]
            for source_base, entries in rotary_alt_entries_map.items()
        }
        all_alt_map = {
            source_base: [entry['part_number'] for entry in entries]
            for source_base, entries in all_alt_entries_map.items()
        }

        customer_id = _get_marketplace_customer_id()
        parts_payload = [
            {'part_number': row['part_number'] or row['base_part_number'], 'quantity': 1}
            for row in rows
        ]
        estimates = _get_portal_estimates(parts_payload, customer_id)
        rotary_alt_stock_rollup_map = _get_alt_stock_rollup_map(customer_id, rotary_alt_entries_map)
        all_alt_stock_rollup_map = _get_alt_stock_rollup_map(customer_id, all_alt_entries_map)

        default_lead_days = _coerce_int(get_portal_setting('default_lead_time_days', 7), default=7)

        # Calculate prices based on portal estimates
        parts_data = []
        for row in rows:
            estimate = estimates.get(row['base_part_number']) or estimates.get(row['part_number'])
            primary_stock_qty = estimate.get('stock_quantity') if estimate else None
            selected_alt_rollup = _select_alt_stock_rollup(
                include_non_hqpl_alts,
                rotary_alt_stock_rollup_map.get(row['base_part_number']),
                all_alt_stock_rollup_map.get(row['base_part_number']),
            )
            effective_estimate = (
                _apply_alt_stock_rollup_to_estimate(estimate, selected_alt_rollup)
                if include_alt_stock_rollup
                else (estimate or {})
            )
            price = effective_estimate.get('estimated_price')
            lead_time = effective_estimate.get('estimated_lead_days')
            in_stock = bool(effective_estimate.get('in_stock'))
            stock_qty = effective_estimate.get('stock_quantity')

            quantity = default_quantity
            stock_qty_value = _coerce_int(stock_qty, default=0)
            if stock_qty_value > 0:
                quantity = stock_qty_value

            if lead_time is None:
                lead_time = 1 if in_stock else default_lead_days
            lead_time_value = _sanitize_marketplace_lead_time_days(
                lead_time,
                default=default_lead_days,
            )

            part_number = row['part_number'] or row['base_part_number']
            resolved_description = _coerce_text(row['mkp_description'], default='')
            if not resolved_description and export_defaults['description_mode'] == 'part_number':
                resolved_description = part_number or ''
            alt_stock_note = (
                _build_alt_stock_note(primary_stock_qty, selected_alt_rollup)
                if include_alt_stock_rollup
                else ''
            )
            if alt_stock_note:
                resolved_description = ' '.join(
                    segment for segment in [resolved_description, alt_stock_note] if segment
                )
            baseline_row = (
                baseline_rows.get(row['base_part_number'])
                or baseline_rows.get(part_number)
                or None
            )

            parts_data.append({
                'base_part_number': row['base_part_number'],
                'part_number': part_number,
                'mkp_category': row['mkp_category'],
                'mkp_description': resolved_description,
                'mkp_name': row['mkp_name'],
                'mkp_product_summary': row['mkp_product_summary'],
                'mkp_product_presentation': row['mkp_product_presentation'],
                'mkp_product_unit': _coerce_text(row['mkp_product_unit'], default=export_defaults['mkp_product_unit'] or ''),
                'mkp_package_content': _coerce_int(row['mkp_package_content'], default=export_defaults['mkp_package_content']),
                'mkp_package_content_unit': _coerce_text(row['mkp_package_content_unit'], default=export_defaults['mkp_package_content_unit'] or ''),
                'mkp_third_level': _coerce_text(row['mkp_third_level'], default=export_defaults['mkp_third_level'] or ''),
                'mkp_dangerous': _coerce_bool(row['mkp_dangerous'], default=export_defaults['mkp_dangerous']),
                'mkp_eccn': _coerce_text(row['mkp_eccn'], default=export_defaults['mkp_eccn'] or ''),
                'mkp_serialized': _coerce_bool(row['mkp_serialized'], default=export_defaults['mkp_serialized']),
                'mkp_log_card': _coerce_bool(row['mkp_log_card'], default=export_defaults['mkp_log_card']),
                'mkp_easaf1': _coerce_bool(row['mkp_easaf1'], default=export_defaults['mkp_easaf1']),
                'mkp_offer_product_id': _coerce_text(row['mkp_offer_product_id'], default=''),
                'mkp_offer_product_id_type': _coerce_text(row['mkp_offer_product_id_type'], default='SKU'),
                'include_non_hqpl_alts': include_non_hqpl_alts,
                'include_alt_stock_rollup': include_alt_stock_rollup,
                'all_global_alt_part_numbers': all_alt_map.get(row['base_part_number'], []),
                'rotary_hqpl_alt_part_numbers': rotary_alt_map.get(row['base_part_number'], []),
                'selected_alt_stock_rollup': selected_alt_rollup,
                'description': resolved_description,
                'manufacturer': _coerce_text(row['manufacturer_name'], default=''),
                'quantity': quantity,
                'source_cost': price,
                'source_currency': effective_estimate.get('currency') if effective_estimate else None,
                'price_source': effective_estimate.get('price_source') if effective_estimate else None,
                'pricing_debug': effective_estimate.get('debug_info') if effective_estimate else None,
                'source_lead_days': effective_estimate.get('estimated_lead_days') if effective_estimate else None,
                'source_in_stock': in_stock,
                'source_stock_qty': stock_qty_value,
                'price': _convert_marketplace_price_to_eur(price) if price is not None else 0.0,
                'condition': 'New',
                'lead_time_days': lead_time_value,
                'baseline_row': baseline_row,
                **_resolve_airbus_mpn_title(part_number),
            })

        if export_mode == 'offers':
            csv_rows = []
            debug_rows = []
            for part in parts_data:
                part_number = _part_identifier(part)
                description = _coerce_text(part.get('mkp_description') or part.get('description'), default=part_number)
                offer_product_id, offer_product_id_type = _resolve_offer_identity(part, source_mode)
                offer_row = _build_offer_row_from_payload({
                    'sku': part_number,
                    'product-id': offer_product_id,
                    'product-id-type': offer_product_id_type,
                    'description': description,
                    'internal-description': '',
                    'price': part.get('price'),
                    'price-additional-info': '',
                    'quantity': part.get('quantity'),
                    'min-quantity-alert': '',
                    'state': '1',
                    'available-start-date': '',
                    'available-end-date': '',
                    'logistic-class': '',
                    'favorite-rank': '',
                    'discount-start-date': '',
                    'discount-end-date': '',
                    'discount-price': '',
                    'update-delete': 'update',
                    'allow-quote-requests': 'true',
                    'leadtime-to-ship': part.get('lead_time_days'),
                    'min-order-quantity': '',
                    'max-order-quantity': '',
                    'package-quantity': '',
                    'commercial-on-collection': 'ON_DEMAND',
                    'plt': '',
                    'plt-unit': '',
                    'shelflife': '',
                    'shelflife-unit': '',
                    'warranty': '',
                    'warranty-unit': '',
                    'up-sell': '',
                    'cross-sell': '',
                    'vendor-reference': '',
                    'baseline_row': part.get('baseline_row'),
                })
                csv_rows.append(offer_row)

                if debug_offer_export:
                    debug_rows.append(_build_offer_debug_row(part, offer_row, default_quantity))

            skipped_invalid = []
            if skip_invalid_mandatory:
                csv_rows, skipped_invalid = _split_valid_rows(
                    csv_rows,
                    _get_offer_missing_required_fields,
                    lambda row: str(row.get('sku') or '').strip(),
                )
                if debug_offer_export:
                    valid_skus = {str(row.get('sku') or '').strip() for row in csv_rows}
                    debug_rows = [
                        row for row in debug_rows
                        if str(row.get('sku') or '').strip() in valid_skus
                    ]
                if not csv_rows:
                    return jsonify({
                        'error': 'All selected lines are missing mandatory offer fields.',
                        'skipped_invalid_count': len(skipped_invalid),
                        'skipped_invalid_preview': skipped_invalid[:20],
                    }), 400

            if debug_offer_export:
                fieldnames = [
                    'sku',
                    'product-id',
                    'product-id-type',
                    'winning_source',
                    'source_type',
                    'source_reference',
                    'source_supplier',
                    'source_date',
                    'price_source',
                    'source_currency',
                    'source_cost',
                    'source_cost_in_gbp',
                    'margin_pct',
                    'target_price_gbp',
                    'rounded_price_gbp',
                    'converted_price_eur',
                    'exported_offer_price',
                    'in_stock',
                    'stock_qty',
                    'default_quantity_input',
                    'exported_offer_quantity',
                    'source_lead_days',
                    'exported_lead_days',
                    'pricing_pathway',
                ]
                csv_text = io.StringIO()
                writer = csv.DictWriter(csv_text, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(debug_rows)
                csv_file = io.BytesIO(csv_text.getvalue().encode('utf-8'))
                csv_file.seek(0)
                filename_prefix = "AH_Marketplace_Offers_Debug"
            else:
                csv_bytes = build_offers_csv(csv_rows, validate_required=skip_invalid_mandatory)
                csv_file = io.BytesIO(csv_bytes)
                csv_file.seek(0)
                filename_prefix = "AH_Marketplace_Offers"
        else:
            skipped_invalid = []
            if skip_invalid_mandatory:
                parts_data, skipped_invalid = _split_valid_rows(
                    parts_data,
                    _get_product_missing_required_fields,
                    _part_identifier,
                )
                if not parts_data:
                    return jsonify({
                        'error': 'All selected lines are missing mandatory fields.',
                        'skipped_invalid_count': len(skipped_invalid),
                        'skipped_invalid_preview': skipped_invalid[:20],
                    }), 400

            csv_file = export_parts_to_airbus_marketplace_csv(parts_data)
            filename_prefix = "AH_Marketplace_Upload"

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{filename_prefix}_{timestamp}.csv"

        return send_file(
            csv_file,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.exception("Error exporting to marketplace")
        return jsonify({'error': str(e)}), 500


@marketplace_bp.route('/auto-categorize', methods=['POST'])
def auto_categorize_parts():
    """
    Auto-categorize parts in batches with progress tracking

    POST body:
    {
        "overwrite": false,
        "base_part_numbers": [],
        "batch_size": 10,  // Process this many at a time
        "offset": 0  // Start from this position
    }
    """
    try:
        data = request.get_json() or {}
        overwrite = data.get('overwrite', False)
        base_part_numbers = data.get('base_part_numbers', [])
        batch_size = data.get('batch_size', 10)
        offset = data.get('offset', 0)

        db = get_db()
        cursor = db.cursor()

        # Build query
        if base_part_numbers:
            placeholders = ','.join('?' * len(base_part_numbers))
            query = f"""
                SELECT base_part_number, part_number
                FROM part_numbers 
                WHERE base_part_number IN ({placeholders})
            """
            params = base_part_numbers
        else:
            query = """
                SELECT base_part_number, part_number
                FROM part_numbers
            """
            if not overwrite:
                query += " WHERE mkp_category IS NULL OR mkp_category = ''"
            params = []

        query += f" LIMIT {batch_size} OFFSET {offset}"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        if not rows:
            return jsonify({
                'success': True,
                'message': 'All parts processed',
                'updated_count': 0,
                'total_in_batch': 0,
                'has_more': False
            }), 200

        # Process this batch
        parts_list = [
            {
                'part_number': row['part_number'] or row['base_part_number'],
                'description': '',
                'additional_info': ''
            }
            for row in rows
        ]

        suggestions = suggest_categories_batch(parts_list)

        # Update database
        updated_count = 0
        batch_details = []
        for row, suggestion in zip(rows, suggestions):
            base_part_number = row['base_part_number']
            suggested_category = suggestion.get('suggested_category')
            was_updated = False

            if suggested_category:
                cursor.execute(
                    "UPDATE part_numbers SET mkp_category = ? WHERE base_part_number = ?",
                    (suggested_category, base_part_number)
                )
                updated_count += 1
                was_updated = True

            batch_details.append({
                'base_part_number': base_part_number,
                'part_number': row['part_number'],
                'suggested_category': suggested_category,
                'confidence': suggestion.get('confidence'),
                'reasoning': suggestion.get('reasoning'),
                'updated': was_updated
            })

        db.commit()

        # Check if there are more parts to process
        cursor.execute(query.replace(f"LIMIT {batch_size} OFFSET {offset}", f"LIMIT 1 OFFSET {offset + batch_size}"), params)
        has_more = cursor.fetchone() is not None

        return jsonify({
            'success': True,
            'message': f'Processed batch of {len(rows)} parts',
            'updated_count': updated_count,
            'total_in_batch': len(rows),
            'has_more': has_more,
            'next_offset': offset + batch_size,
            'batch_details': batch_details
        }), 200

    except Exception as e:
        logger.exception("Error auto-categorizing parts")
        if db:
            db.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@marketplace_bp.route('/mirakl/health', methods=['GET'])
def mirakl_health():
    client, error = _get_mirakl_client()
    env_base_url = os.getenv('MIRAKL_BASE_URL')
    env_shop_id = os.getenv('MIRAKL_SHOP_ID')
    env_api_key = os.getenv('MIRAKL_API_KEY')
    portal_base_url = get_portal_setting('mirakl_base_url')
    portal_shop_id = get_portal_setting('mirakl_shop_id')
    portal_api_key = get_portal_setting('mirakl_api_key')
    base_url = env_base_url or portal_base_url
    shop_id = env_shop_id or portal_shop_id
    api_key = env_api_key or portal_api_key
    diagnostics = {
        'base_url': base_url,
        'shop_id': shop_id,
        'api_key_masked': _mask_api_key(api_key),
        'api_key_length': len(api_key) if api_key else 0,
        'sources': {
            'base_url': 'env' if env_base_url else ('portal' if portal_base_url else None),
            'shop_id': 'env' if env_shop_id else ('portal' if portal_shop_id else None),
            'api_key': 'env' if env_api_key else ('portal' if portal_api_key else None),
        },
        'headers': ['Authorization'] + (['X-Mirakl-Shop-Id'] if shop_id else []),
    }
    if error:
        return jsonify({'success': False, 'error': error, 'diagnostics': diagnostics}), 400

    try:
        data = client.get_account()
        return jsonify({'success': True, 'account': data, 'diagnostics': diagnostics}), 200
    except MiraklError as exc:
        logger.exception("Mirakl health check failed")
        return jsonify({'success': False, 'error': str(exc), 'diagnostics': diagnostics}), 502


@marketplace_bp.route('/mirakl/offers/import', methods=['POST'])
def mirakl_import_offers():
    client, error = _get_mirakl_client()
    if error:
        return jsonify({'success': False, 'error': error}), 400

    data = request.get_json() or {}
    offers = data.get('offers') or []
    import_mode = data.get('import_mode', 'NORMAL')
    skip_invalid_mandatory = _coerce_bool(data.get('skip_invalid_mandatory'), default=False)

    if not offers:
        return jsonify({'success': False, 'error': 'offers array is required'}), 400

    skipped_invalid = []
    offers = [_build_offer_row_from_payload(offer) for offer in offers]
    for offer in offers:
        offer['price'] = _convert_marketplace_price_to_eur(offer.get('price'))
    if skip_invalid_mandatory:
        offers, skipped_invalid = _split_valid_rows(
            offers,
            _get_offer_missing_required_fields,
            lambda row: str(row.get('sku') or '').strip(),
        )
        if not offers:
            return jsonify({
                'success': False,
                'error': 'All provided offers are missing mandatory fields.',
                'skipped_invalid_count': len(skipped_invalid),
                'skipped_invalid_preview': skipped_invalid[:20],
            }), 400

    try:
        csv_bytes = build_offers_csv(offers)
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400

    try:
        result = client.import_offers(csv_bytes, import_mode=import_mode)
        logger.info("Mirakl offer import result: %s", result)
        stored_offer_product_ids_count = _store_offer_identity_updates_from_rows(offers)
        return jsonify({
            'success': True,
            'result': result,
            'stored_offer_product_ids_count': stored_offer_product_ids_count,
            'skipped_invalid_count': len(skipped_invalid),
            'skipped_invalid_preview': skipped_invalid[:20],
        }), 200
    except MiraklError as exc:
        logger.exception("Mirakl offer import failed")
        return jsonify({'success': False, 'error': str(exc)}), 502


@marketplace_bp.route('/master-list-test-offers', methods=['GET'])
def download_master_list_test_offers():
    try:
        offers, skipped = _build_master_list_test_offers()
        csv_bytes = build_offers_csv(offers)
        csv_file = io.BytesIO(csv_bytes)
        csv_file.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"AH_Master_List_Test_Offers_{timestamp}.csv"
        response = send_file(
            csv_file,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
        if skipped:
            response.headers['X-Master-List-Test-Skipped'] = str(len(skipped))
        return response
    except Exception as exc:
        logger.exception("Error building master-list test offers CSV")
        return jsonify({'success': False, 'error': str(exc)}), 500


@marketplace_bp.route('/mirakl/products/import', methods=['POST'])
def mirakl_import_products():
    """
    Import products using the full Airbus template format (products + offers combined).
    This is the correct endpoint for the Airbus marketplace.
    """
    client, error = _get_mirakl_client()
    if error:
        return jsonify({'success': False, 'error': error}), 400

    data = request.get_json() or {}
    parts = data.get('parts') or []
    import_mode = data.get('import_mode', 'NORMAL')
    skip_invalid_mandatory = _coerce_bool(data.get('skip_invalid_mandatory'), default=False)

    if not parts:
        return jsonify({'success': False, 'error': 'parts array is required'}), 400

    skipped_invalid = []
    for part in parts:
        part['price'] = _convert_marketplace_price_to_eur(part.get('price'))
    if skip_invalid_mandatory:
        parts, skipped_invalid = _split_valid_rows(
            parts,
            _get_product_missing_required_fields,
            _part_identifier,
        )
        if not parts:
            return jsonify({
                'success': False,
                'error': 'All provided product lines are missing mandatory fields.',
                'skipped_invalid_count': len(skipped_invalid),
                'skipped_invalid_preview': skipped_invalid[:20],
            }), 400

    try:
        # Use the Airbus marketplace export function to build the full template CSV
        csv_file = export_parts_to_airbus_marketplace_csv(parts)
        csv_bytes = csv_file.read()
        logger.info("Built Airbus template CSV with %d parts (%d bytes)", len(parts), len(csv_bytes))
    except Exception as exc:
        logger.exception("Failed to build Airbus template CSV")
        return jsonify({'success': False, 'error': str(exc)}), 400

    try:
        result = client.import_products(csv_bytes, import_mode=import_mode)
        logger.info("Mirakl product import result: %s", result)
        return jsonify({
            'success': True,
            'result': result,
            'skipped_invalid_count': len(skipped_invalid),
            'skipped_invalid_preview': skipped_invalid[:20],
        }), 200
    except MiraklError as exc:
        logger.exception("Mirakl product import failed")
        return jsonify({'success': False, 'error': str(exc)}), 502


@marketplace_bp.route('/mirakl/offers/import-file', methods=['POST'])
def mirakl_import_offers_file():
    client, error = _get_mirakl_client()
    if error:
        return jsonify({'success': False, 'error': error}), 400

    file = request.files.get('file')
    import_mode = (request.form.get('import_mode') or 'NORMAL').strip() or 'NORMAL'

    if not file or not file.filename:
        return jsonify({'success': False, 'error': 'CSV file is required'}), 400

    if not file.filename.lower().endswith('.csv'):
        return jsonify({'success': False, 'error': 'Only .csv files are supported'}), 400

    csv_bytes = file.read()
    if not csv_bytes:
        return jsonify({'success': False, 'error': 'CSV file is empty'}), 400
    csv_bytes = _dedupe_csv_headers(csv_bytes)

    try:
        result = client.import_offers(csv_bytes, import_mode=import_mode)
        logger.info("Mirakl offer import (file) result: %s", result)
        return jsonify({'success': True, 'result': result}), 200
    except MiraklError as exc:
        logger.exception("Mirakl offer import (file) failed")
        return jsonify({'success': False, 'error': str(exc)}), 502


@marketplace_bp.route('/mirakl/products/import-file', methods=['POST'])
def mirakl_import_products_file():
    client, error = _get_mirakl_client()
    if error:
        return jsonify({'success': False, 'error': error}), 400

    file = request.files.get('file')
    import_mode = (request.form.get('import_mode') or 'NORMAL').strip() or 'NORMAL'

    if not file or not file.filename:
        return jsonify({'success': False, 'error': 'CSV file is required'}), 400

    if not file.filename.lower().endswith('.csv'):
        return jsonify({'success': False, 'error': 'Only .csv files are supported'}), 400

    csv_bytes = file.read()
    if not csv_bytes:
        return jsonify({'success': False, 'error': 'CSV file is empty'}), 400

    try:
        result = client.import_products(csv_bytes, import_mode=import_mode)
        logger.info("Mirakl product import result: %s", result)
        return jsonify({'success': True, 'result': result}), 200
    except MiraklError as exc:
        logger.exception("Mirakl product import (file) failed")
        return jsonify({'success': False, 'error': str(exc)}), 502


@marketplace_bp.route('/mirakl/offers/imports/<import_id>', methods=['GET'])
def mirakl_import_status(import_id):
    client, error = _get_mirakl_client()
    if error:
        return jsonify({'success': False, 'error': error}), 400

    try:
        result = client.get_offers_import(import_id)
        return jsonify({'success': True, 'result': result}), 200
    except MiraklError as exc:
        logger.exception("Mirakl import status fetch failed")
        return jsonify({'success': False, 'error': str(exc)}), 502


@marketplace_bp.route('/mirakl/offers/imports/<import_id>/errors', methods=['GET'])
def mirakl_import_errors(import_id):
    client, error = _get_mirakl_client()
    if error:
        return jsonify({'success': False, 'error': error}), 400

    try:
        content, content_type = client.get_offers_import_errors(import_id)
        return content, 200, {'Content-Type': content_type}
    except MiraklError as exc:
        logger.exception("Mirakl import error report fetch failed")
        return jsonify({'success': False, 'error': str(exc)}), 502


@marketplace_bp.route('/mirakl/settings', methods=['POST'])
def mirakl_update_settings():
    data = request.get_json() or {}
    base_url = (data.get('mirakl_base_url') or '').strip()
    shop_id = (data.get('mirakl_shop_id') or '').strip()
    api_key = (data.get('mirakl_api_key') or '').strip()

    if not base_url:
        return jsonify({'success': False, 'error': 'Mirakl base URL is required.'}), 400

    db = get_db()
    cursor = None
    try:
        cursor = db.cursor()
        _upsert_portal_setting(cursor, 'mirakl_base_url', base_url)
        _upsert_portal_setting(cursor, 'mirakl_shop_id', shop_id)
        if api_key:
            _upsert_portal_setting(cursor, 'mirakl_api_key', api_key)
        db.commit()
    except Exception as exc:
        logger.exception("Failed to update Mirakl settings")
        db.rollback()
        return jsonify({'success': False, 'error': str(exc)}), 500
    finally:
        if cursor:
            cursor.close()

    return jsonify({'success': True}), 200
