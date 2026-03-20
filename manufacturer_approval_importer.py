import json
import os
import re
from datetime import date, datetime, timezone
from typing import Any, Dict, Iterator, List, Optional, Sequence, Tuple


LIST_TYPES = {
    'airbus_fixed_wing': 'Airbus Fixed Wing',
    'airbus_rotary': 'Airbus Rotary',
}

RAW_COLUMN_MAP = {
    'Manufacturer': 'manufacturer_code',
    'Manufacturer Text': 'manufacturer_name',
    'Location': 'location',
    'Country': 'country',
    'CAGE Co': 'cage_code',
    'Approval Status Text': 'approval_status',
    'Type of Data Text': 'data_type',
    'Standard': 'standard',
    'Airbus Material': 'airbus_material',
    'Airbus Material Text': 'airbus_material_text',
    'Interchangeability Flag': 'interchangeability_flag',
    'Manufacturer PN': 'manufacturer_part_number',
    'Usage Restriction Text': 'usage_restriction',
    'P Status': 'p_status',
    'P Status Text': 'p_status_text',
    'Change date Status P': 'status_change_date',
    'Counter of QIR': 'qir_count',
    'Manufacturer Part Number (MPN)': 'manufacturer_part_number',
    'Manufacturer name': 'manufacturer_name',
    'AH Manufacturer code': 'manufacturer_code',
}

UPSERT_COLUMNS: Sequence[str] = (
    'import_id',
    'approval_list_type',
    'manufacturer_code',
    'manufacturer_name',
    'location',
    'country',
    'cage_code',
    'approval_status',
    'data_type',
    'standard',
    'airbus_material',
    'airbus_material_text',
    'interchangeability_flag',
    'manufacturer_part_number',
    'usage_restriction',
    'p_status',
    'p_status_text',
    'status_change_date',
    'qir_count',
    'airbus_material_base',
    'manufacturer_part_number_base',
    'created_at',
    'updated_at',
)


def clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip() or None
    text = str(value).strip()
    return text or None


def normalize_header(value: Any) -> str:
    return str(value).replace('\n', ' ').strip()


def normalize_part_number(value: Any) -> Optional[str]:
    text = clean_text(value)
    if not text:
        return None
    normalized = re.sub(r'[^A-Z0-9]', '', text.upper())
    return normalized or None


def normalize_country(value: Any) -> Optional[str]:
    text = clean_text(value)
    return text.upper() if text else None


def normalize_cage(value: Any) -> Optional[str]:
    text = clean_text(value)
    return text.upper() if text else None


def parse_status_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    try:
        if len(text) == 8 and text.isdigit():
            return datetime.strptime(text, '%Y%m%d').date()
        return datetime.fromisoformat(text).date()
    except (TypeError, ValueError):
        return None


def parse_int(value: Any) -> Optional[int]:
    text = clean_text(value)
    if text is None:
        return None
    try:
        return int(text)
    except (TypeError, ValueError):
        return None


def iter_records(xlsx_path: str, limit: Optional[int] = None) -> Iterator[Dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [normalize_header(col) for col in header_row]

        yielded = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_map = dict(zip(headers, row))
            normalized: Dict[str, Any] = {}
            for raw_key, normalized_key in RAW_COLUMN_MAP.items():
                normalized[normalized_key] = row_map.get(raw_key)
            yield normalized
            yielded += 1
            if limit and yielded >= limit:
                break
    finally:
        workbook.close()


def build_payload(
    record: Dict[str, Any],
    import_id: int,
    approval_list_type: str,
) -> Tuple[Optional[Tuple[Any, ...]], Optional[str]]:
    manufacturer_name = clean_text(record.get('manufacturer_name'))
    airbus_material = clean_text(record.get('airbus_material'))
    manufacturer_part_number = clean_text(record.get('manufacturer_part_number'))

    if not manufacturer_name:
        return None, 'Missing manufacturer name'
    if not airbus_material and not manufacturer_part_number:
        return None, 'Missing both Airbus Material and Manufacturer PN'

    airbus_material_base = normalize_part_number(airbus_material)
    manufacturer_part_number_base = normalize_part_number(manufacturer_part_number)
    now = datetime.now(tz=timezone.utc)

    return (
        import_id,
        approval_list_type,
        clean_text(record.get('manufacturer_code')),
        manufacturer_name,
        clean_text(record.get('location')),
        normalize_country(record.get('country')),
        normalize_cage(record.get('cage_code')),
        clean_text(record.get('approval_status')),
        clean_text(record.get('data_type')),
        clean_text(record.get('standard')),
        airbus_material,
        clean_text(record.get('airbus_material_text')),
        clean_text(record.get('interchangeability_flag')),
        manufacturer_part_number,
        clean_text(record.get('usage_restriction')),
        clean_text(record.get('p_status')),
        clean_text(record.get('p_status_text')),
        parse_status_date(record.get('status_change_date')),
        parse_int(record.get('qir_count')),
        airbus_material_base,
        manufacturer_part_number_base,
        now,
        now,
    ), None


def start_import(
    cursor,
    source_files: Sequence[str],
    approval_list_type: str,
    imported_by: Optional[str] = None,
    overwrite_existing: bool = True,
) -> int:
    cursor.execute(
        """
        INSERT INTO manufacturer_approval_imports (
            source_file,
            source_files_json,
            source_file_count,
            imported_by,
            approval_list_type,
            overwrite_existing
        )
        VALUES (%s, %s, %s, %s, %s, %s)
        RETURNING id
        """,
        (
            ', '.join(os.path.basename(path) for path in source_files),
            json.dumps(list(source_files)),
            len(source_files),
            imported_by,
            approval_list_type,
            overwrite_existing,
        ),
    )
    return cursor.fetchone()[0]


def delete_existing_rows(cursor, approval_list_type: str) -> int:
    cursor.execute(
        'DELETE FROM manufacturer_approvals WHERE approval_list_type = %s',
        (approval_list_type,),
    )
    return cursor.rowcount or 0


def update_import_row_count(cursor, import_id: int) -> int:
    cursor.execute(
        'SELECT COUNT(*) FROM manufacturer_approvals WHERE import_id = %s',
        (import_id,),
    )
    row_count = cursor.fetchone()[0]
    cursor.execute(
        'UPDATE manufacturer_approval_imports SET row_count = %s WHERE id = %s',
        (row_count, import_id),
    )
    return row_count


def upsert_batch(cursor, rows: List[Tuple[Any, ...]]) -> None:
    from psycopg2.extras import execute_values

    if not rows:
        return

    template = '(' + ', '.join(['%s'] * len(UPSERT_COLUMNS)) + ')'
    query = f"""
        INSERT INTO manufacturer_approvals ({', '.join(UPSERT_COLUMNS)})
        VALUES %s
        ON CONFLICT (
            approval_list_type,
            airbus_material,
            manufacturer_part_number,
            manufacturer_name,
            cage_code,
            location
        )
        DO UPDATE SET
            import_id = EXCLUDED.import_id,
            manufacturer_code = COALESCE(EXCLUDED.manufacturer_code, manufacturer_approvals.manufacturer_code),
            country = COALESCE(EXCLUDED.country, manufacturer_approvals.country),
            approval_status = COALESCE(EXCLUDED.approval_status, manufacturer_approvals.approval_status),
            data_type = COALESCE(EXCLUDED.data_type, manufacturer_approvals.data_type),
            standard = COALESCE(EXCLUDED.standard, manufacturer_approvals.standard),
            airbus_material = COALESCE(EXCLUDED.airbus_material, manufacturer_approvals.airbus_material),
            airbus_material_text = COALESCE(EXCLUDED.airbus_material_text, manufacturer_approvals.airbus_material_text),
            interchangeability_flag = COALESCE(EXCLUDED.interchangeability_flag, manufacturer_approvals.interchangeability_flag),
            manufacturer_part_number = COALESCE(EXCLUDED.manufacturer_part_number, manufacturer_approvals.manufacturer_part_number),
            usage_restriction = COALESCE(EXCLUDED.usage_restriction, manufacturer_approvals.usage_restriction),
            p_status = COALESCE(EXCLUDED.p_status, manufacturer_approvals.p_status),
            p_status_text = COALESCE(EXCLUDED.p_status_text, manufacturer_approvals.p_status_text),
            status_change_date = COALESCE(EXCLUDED.status_change_date, manufacturer_approvals.status_change_date),
            qir_count = COALESCE(EXCLUDED.qir_count, manufacturer_approvals.qir_count),
            airbus_material_base = COALESCE(EXCLUDED.airbus_material_base, manufacturer_approvals.airbus_material_base),
            manufacturer_part_number_base = COALESCE(EXCLUDED.manufacturer_part_number_base, manufacturer_approvals.manufacturer_part_number_base),
            updated_at = EXCLUDED.updated_at
    """
    execute_values(cursor, query, rows, template=template, page_size=len(rows))


def process_workbooks(
    connection,
    workbook_paths: Sequence[str],
    approval_list_type: str,
    *,
    batch_size: int = 2000,
    limit: Optional[int] = None,
    overwrite_existing: bool = True,
    imported_by: Optional[str] = None,
    dry_run: bool = False,
) -> Dict[str, Any]:
    if approval_list_type not in LIST_TYPES:
        raise ValueError(f'Unsupported approval list type: {approval_list_type}')
    if not workbook_paths:
        raise ValueError('At least one workbook is required for import.')

    cursor = connection.cursor()
    import_id = start_import(
        cursor,
        source_files=workbook_paths,
        approval_list_type=approval_list_type,
        imported_by=imported_by,
        overwrite_existing=overwrite_existing,
    )

    deleted_previous_rows = 0
    if overwrite_existing:
        deleted_previous_rows = delete_existing_rows(cursor, approval_list_type)

    stats: Dict[str, Any] = {
        'import_id': import_id,
        'approval_list_type': approval_list_type,
        'files_processed': len(workbook_paths),
        'rows_seen': 0,
        'rows_written': 0,
        'rows_skipped': 0,
        'deleted_previous_rows': deleted_previous_rows,
        'source_files': [os.path.basename(path) for path in workbook_paths],
    }

    batch: List[Tuple[Any, ...]] = []
    for workbook_path in workbook_paths:
        for record in iter_records(workbook_path, limit=limit):
            stats['rows_seen'] += 1
            payload, skip_reason = build_payload(record, import_id, approval_list_type)
            if payload is None:
                stats['rows_skipped'] += 1
                continue
            batch.append(payload)
            if len(batch) >= batch_size:
                upsert_batch(cursor, batch)
                batch.clear()

    if batch:
        upsert_batch(cursor, batch)

    stats['rows_written'] = update_import_row_count(cursor, import_id)

    if dry_run:
        connection.rollback()
    else:
        connection.commit()

    return stats
