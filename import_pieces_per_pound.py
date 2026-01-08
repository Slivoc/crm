"""
Import pieces per pound data from CSV/XLSX into PostgreSQL.

This script reads part numbers and their PPP values, normalizes the part numbers
to base_part_number format, and updates or inserts parts in the database.

Expected CSV format:
    part_number,pieces_per_pound
    MS35338-44,1250
    NAS6606D46,850

Usage:
    python import_pieces_per_pound.py ppp_data.csv
    python import_pieces_per_pound.py ppp_data.xlsx
"""
import argparse
import csv
import logging
import os
import re
from typing import Any, Dict, Iterator, List, Optional, Tuple

from dotenv import load_dotenv
import psycopg2
from psycopg2.extras import execute_values

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import pieces per pound data into PostgreSQL."
    )
    parser.add_argument("file_path", help="Path to the CSV or XLSX file.")
    parser.add_argument(
        "--batch-size",
        type=int,
        default=1000,
        help="Number of rows to upsert per batch (default: 1000).",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limit number of data rows for testing.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Run the parser and SQL without committing changes.",
    )
    return parser.parse_args()


def normalize_part_number(part_number: Any) -> Optional[str]:
    """
    Normalize part number by removing special characters and converting to uppercase.
    This matches the base_part_number format in the database.
    """
    if not part_number:
        return None

    text = str(part_number).strip()
    if not text:
        return None

    # Remove all special characters (dashes, spaces, slashes, etc) and convert to uppercase
    normalized = re.sub(r'[^A-Z0-9]', '', text.upper())
    return normalized if normalized else None


def parse_decimal(value: Any) -> Optional[float]:
    """Parse a decimal/float value."""
    if value is None or value == '':
        return None

    try:
        return float(value)
    except (ValueError, TypeError):
        logger.warning(f"Could not parse decimal value: {value}")
        return None


def iter_csv_records(csv_path: str, limit: Optional[int]) -> Iterator[Dict[str, Any]]:
    """Iterate over CSV records."""
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            if limit and i >= limit:
                break
            yield row


def iter_xlsx_records(xlsx_path: str, limit: Optional[int]) -> Iterator[Dict[str, Any]]:
    """Iterate over XLSX records."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required to read XLSX files. Install with: pip install openpyxl")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(col).strip() if col else f"Column_{i}" for i, col in enumerate(header_row)]

        records_yielded = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_map = dict(zip(headers, row))
            yield row_map
            records_yielded += 1
            if limit and records_yielded >= limit:
                break
    finally:
        workbook.close()


def iter_records(file_path: str, limit: Optional[int]) -> Iterator[Dict[str, Any]]:
    """Iterate over records from CSV or XLSX file."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.csv':
        return iter_csv_records(file_path, limit)
    elif ext in ('.xlsx', '.xls'):
        return iter_xlsx_records(file_path, limit)
    else:
        raise ValueError(f"Unsupported file format: {ext}. Use .csv or .xlsx")


def build_payload(record: Dict[str, Any]) -> Optional[Tuple[str, float, str, str]]:
    """
    Build a payload for insertion/update.
    Returns (base_part_number, pieces_per_pound, original_part_number, skip_reason) or None.
    """
    # Try common column names
    part_number = (
        record.get('part_number') or
        record.get('Part Number') or
        record.get('PartNumber') or
        record.get('part') or
        record.get('Part')
    )

    ppp_value = (
        record.get('pieces_per_pound') or
        record.get('Pieces Per Pound') or
        record.get('PiecesPerPound') or
        record.get('ppp') or
        record.get('PPP') or
        record.get('pieces_pound')
    )

    if not part_number:
        return None

    original_pn = str(part_number).strip()
    base_part_number = normalize_part_number(part_number)

    if not base_part_number:
        logger.warning(f"Could not normalize part number: {original_pn}")
        return None

    ppp = parse_decimal(ppp_value)
    if ppp is None or ppp <= 0:
        logger.warning(f"Invalid PPP value for {original_pn}: {ppp_value}")
        return None

    return (base_part_number, ppp, original_pn, None)


def upsert_batch(cursor, rows: List[Tuple[str, float, str]]) -> Tuple[int, int]:
    """
    Upsert a batch of PPP records.
    Returns (updated_count, inserted_count).
    """
    if not rows:
        return (0, 0)

    # First, update existing parts
    update_query = """
        UPDATE part_numbers
        SET pieces_per_pound = data.ppp
        FROM (VALUES %s) AS data(base_pn, ppp, orig_pn)
        WHERE part_numbers.base_part_number = data.base_pn
    """
    execute_values(cursor, update_query, rows, template="(%s, %s, %s)")
    updated_count = cursor.rowcount

    # Then, insert new parts that don't exist
    insert_query = """
        INSERT INTO part_numbers (base_part_number, part_number, pieces_per_pound, created_at)
        SELECT data.base_pn, data.orig_pn, data.ppp, CURRENT_TIMESTAMP
        FROM (VALUES %s) AS data(base_pn, ppp, orig_pn)
        WHERE NOT EXISTS (
            SELECT 1 FROM part_numbers WHERE base_part_number = data.base_pn
        )
    """
    execute_values(cursor, insert_query, rows, template="(%s, %s, %s)")
    inserted_count = cursor.rowcount

    return (updated_count, inserted_count)


def process_file(
    connection,
    file_path: str,
    batch_size: int,
    limit: Optional[int],
    dry_run: bool,
) -> Dict[str, int]:
    """Process the import file."""
    cursor = connection.cursor()

    stats = {
        "rows_seen": 0,
        "rows_updated": 0,
        "rows_inserted": 0,
        "rows_skipped": 0
    }

    batch: List[Tuple[str, float, str]] = []

    logger.info(f"Processing file: {file_path}")

    for record in iter_records(file_path, limit):
        stats["rows_seen"] += 1

        result = build_payload(record)
        if result is None:
            stats["rows_skipped"] += 1
            continue

        base_pn, ppp, orig_pn, _ = result
        batch.append((base_pn, ppp, orig_pn))

        if len(batch) >= batch_size:
            updated, inserted = upsert_batch(cursor, batch)
            stats["rows_updated"] += updated
            stats["rows_inserted"] += inserted
            logger.info(f"Processed {stats['rows_seen']} rows so far...")
            batch.clear()

    # Process remaining rows
    if batch:
        updated, inserted = upsert_batch(cursor, batch)
        stats["rows_updated"] += updated
        stats["rows_inserted"] += inserted

    if dry_run:
        logger.info("Dry-run enabled; rolling back transaction.")
        connection.rollback()
    else:
        connection.commit()
        logger.info("Import committed.")

    logger.info(
        "Import summary: %s rows seen, %s updated, %s inserted, %s skipped",
        stats["rows_seen"],
        stats["rows_updated"],
        stats["rows_inserted"],
        stats["rows_skipped"],
    )

    return stats


def main() -> None:
    args = parse_args()
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    load_dotenv()
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL environment variable is required.")

    connection = psycopg2.connect(database_url)
    try:
        process_file(
            connection,
            file_path=args.file_path,
            batch_size=args.batch_size,
            limit=args.limit,
            dry_run=args.dry_run,
        )
    finally:
        connection.close()


if __name__ == "__main__":
    main()
